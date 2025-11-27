"""
Script to generate Excel file with proper nouns for transcript segmentation.
Run this script to create/update the proper_nouns.xlsx file.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

def generate_proper_nouns_excel():
    """Generate Excel file with categorized proper nouns"""
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proper Nouns"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Category style
    category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    category_font = Font(bold=True, size=11)
    
    # Set column headers
    ws['A1'] = "Category"
    ws['B1'] = "Proper Noun Phrase"
    ws['C1'] = "Notes"
    
    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    
    row = 2
    
    # Define categories and their proper nouns
    categories = {
        "Countries & Regions": [
            "United States", "United Kingdom", "United Nations", "United Arab Emirates",
            "South Korea", "North Korea", "South Africa", "North America", "South America",
            "New Zealand", "New York", "New Jersey", "New Hampshire", "New Mexico",
            "Middle East", "Far East", "East Asia", "Southeast Asia", "West Africa",
            "East Africa", "Central America", "Latin America", "European Union",
            "Great Britain", "Soviet Union", "Russian Federation", "Hong Kong",
            "Sri Lanka", "Costa Rica", "Puerto Rico", "El Salvador", "San Marino",
            "Saudi Arabia", "United States of America", "People's Republic of China",
            "Republic of China", "South Sudan", "North Macedonia", "East Timor"
        ],
        
        "Cities & Places": [
            "Los Angeles", "San Francisco", "San Diego", "San Antonio", "San Jose",
            "New York City", "New Orleans", "New Delhi", "New Haven", "New Brunswick",
            "Las Vegas", "Kuala Lumpur", "Buenos Aires", "Rio de Janeiro", "São Paulo",
            "Mexico City", "Buenos Aires", "Cape Town", "Tel Aviv", "Tel Aviv-Yafo",
            "Saint Petersburg", "Saint Louis", "Saint Paul", "Saint John", "Saint Helens",
            "Mount Everest", "Mount Kilimanjaro", "Mount Fuji", "Mount Rushmore",
            "Grand Canyon", "Great Wall", "Great Barrier Reef", "Great Lakes",
            "Pacific Ocean", "Atlantic Ocean", "Indian Ocean", "Arctic Ocean",
            "Mediterranean Sea", "Red Sea", "Black Sea", "Dead Sea"
        ],
        
        "Historical Figures": [
            "Abraham Lincoln", "George Washington", "Thomas Jefferson", "Benjamin Franklin",
            "Martin Luther King", "Nelson Mandela", "Winston Churchill", "Mahatma Gandhi",
            "Albert Einstein", "Isaac Newton", "Charles Darwin", "Marie Curie",
            "Leonardo da Vinci", "Michelangelo", "Pablo Picasso", "Vincent van Gogh",
            "William Shakespeare", "Jane Austen", "Charles Dickens", "Mark Twain",
            "Confucius", "Socrates", "Plato", "Aristotle", "Galileo Galilei",
            "Napoleon Bonaparte", "Julius Caesar", "Alexander the Great", "Genghis Khan",
            "Queen Elizabeth", "King George", "Princess Diana", "Queen Victoria"
        ],
        
        "Companies & Brands": [
            "Toyota", "Honda", "BMW", "Mercedes-Benz", "Volkswagen", "Ford Motor",
            "General Motors", "Apple Inc", "Microsoft Corporation", "Google LLC",
            "Amazon.com", "Facebook", "Meta Platforms", "Tesla Inc", "SpaceX",
            "Boeing Company", "Airbus", "Samsung Electronics", "Sony Corporation",
            "Nike Inc", "Adidas", "Coca-Cola", "PepsiCo", "McDonald's Corporation",
            "Starbucks Corporation", "Walmart Inc", "Target Corporation", "IKEA",
            "Nintendo", "PlayStation", "Xbox", "Netflix", "Disney", "Warner Bros"
        ],
        
        "Educational Institutions": [
            "Harvard University", "Yale University", "Princeton University", "Stanford University",
            "Massachusetts Institute of Technology", "MIT", "Oxford University", "Cambridge University",
            "University of Cambridge", "University of Oxford", "Columbia University", "University of Chicago",
            "University of Pennsylvania", "Cornell University", "Dartmouth College", "Brown University",
            "University of California", "UC Berkeley", "UCLA", "UC San Diego", "UC Los Angeles",
            "New York University", "NYU", "Boston University", "Boston College", "Georgetown University",
            "Johns Hopkins University", "Duke University", "Northwestern University", "Vanderbilt University",
            "Rice University", "Emory University", "Carnegie Mellon University", "Carnegie Mellon",
            "University of Michigan", "University of Texas", "UT Austin", "Texas A&M University",
            "University of Southern California", "USC", "University of Washington", "UW Seattle"
        ],
        
        "Organizations & Institutions": [
            "World Health Organization", "WHO", "United Nations", "UN", "UNESCO", "UNICEF",
            "World Bank", "International Monetary Fund", "IMF", "World Trade Organization", "WTO",
            "North Atlantic Treaty Organization", "NATO", "European Union", "EU",
            "Red Cross", "Red Crescent", "Amnesty International", "Greenpeace", "World Wildlife Fund",
            "Bill and Melinda Gates Foundation", "Gates Foundation", "Ford Foundation",
            "Rockefeller Foundation", "Carnegie Foundation", "National Science Foundation", "NSF",
            "National Institutes of Health", "NIH", "Centers for Disease Control", "CDC",
            "Federal Bureau of Investigation", "FBI", "Central Intelligence Agency", "CIA",
            "National Aeronautics and Space Administration", "NASA", "European Space Agency", "ESA"
        ],
        
        "Historical Events": [
            "World War I", "World War II", "World War 1", "World War 2", "First World War", "Second World War",
            "Cold War", "Civil War", "Vietnam War", "Korean War", "Gulf War", "Iraq War",
            "American Revolution", "French Revolution", "Russian Revolution", "Industrial Revolution",
            "Renaissance", "Enlightenment", "Reformation", "Crusades", "Black Death",
            "Great Depression", "Great Recession", "Dot-com Bubble", "Financial Crisis",
            "September 11", "9/11", "Pearl Harbor", "D-Day", "Battle of Gettysburg",
            "Fall of Berlin Wall", "Berlin Wall", "Cuban Missile Crisis", "Space Race"
        ],
        
        "Academic Terms & Concepts": [
            "Bachelor of Arts", "Bachelor of Science", "Bachelor of Engineering", "Bachelor of Business",
            "Master of Arts", "Master of Science", "Master of Business Administration", "MBA",
            "Doctor of Philosophy", "PhD", "Doctor of Medicine", "MD", "Doctor of Education", "EdD",
            "Associate Degree", "Bachelor's Degree", "Master's Degree", "Doctoral Degree",
            "Grade Point Average", "GPA", "Scholastic Assessment Test", "SAT", "American College Testing", "ACT",
            "Graduate Record Examination", "GRE", "Graduate Management Admission Test", "GMAT",
            "International English Language Testing System", "IELTS", "Test of English as Foreign Language", "TOEFL"
        ],
        
        "Scientific Concepts & Theories": [
            "Big Bang Theory", "Theory of Relativity", "General Relativity", "Special Relativity",
            "Quantum Mechanics", "Quantum Theory", "String Theory", "Evolution Theory",
            "Natural Selection", "Survival of the Fittest", "Theory of Evolution",
            "Periodic Table", "Periodic Table of Elements", "DNA", "RNA", "Genetic Code",
            "Newton's Laws", "Newton's First Law", "Newton's Second Law", "Newton's Third Law",
            "Pythagorean Theorem", "Einstein's Equation", "E=mc²", "Schrodinger's Cat",
            "Heisenberg Uncertainty Principle", "Planck's Constant", "Avogadro's Number"
        ],
        
        "Government & Politics": [
            "Prime Minister", "President of the United States", "Vice President", "Secretary of State",
            "Secretary of Defense", "Attorney General", "Supreme Court", "Supreme Court Justice",
            "House of Representatives", "House of Commons", "House of Lords", "Senate",
            "United States Congress", "US Congress", "Parliament", "British Parliament",
            "European Parliament", "United Nations General Assembly", "Security Council",
            "White House", "Capitol Hill", "Downing Street", "10 Downing Street",
            "Federal Reserve", "Fed", "Internal Revenue Service", "IRS", "Social Security Administration"
        ],
        
        "Religious & Philosophical": [
            "Roman Catholic Church", "Catholic Church", "Protestant Reformation", "Eastern Orthodox Church",
            "Buddhism", "Christianity", "Islam", "Judaism", "Hinduism", "Sikhism",
            "Bible", "Quran", "Koran", "Torah", "Talmud", "Vedas", "Bhagavad Gita",
            "Jesus Christ", "Muhammad", "Buddha", "Moses", "Abraham", "Noah",
            "Pope Francis", "Pope John Paul", "Dalai Lama", "Archbishop of Canterbury"
        ],
        
        "Technology & Innovation": [
            "Artificial Intelligence", "AI", "Machine Learning", "Deep Learning", "Neural Networks",
            "Internet of Things", "IoT", "Cloud Computing", "Big Data", "Data Science",
            "Blockchain Technology", "Cryptocurrency", "Bitcoin", "Ethereum", "NFT",
            "Virtual Reality", "VR", "Augmented Reality", "AR", "Mixed Reality", "MR",
            "5G Network", "4G Network", "Wi-Fi", "Bluetooth", "USB", "HDMI",
            "Operating System", "Windows", "Mac OS", "Linux", "Android", "iOS"
        ],
        
        "Arts & Literature": [
            "Renaissance Art", "Baroque Period", "Impressionism", "Cubism", "Surrealism",
            "Mona Lisa", "Starry Night", "The Scream", "Guernica", "Sistine Chapel",
            "Hamlet", "Romeo and Juliet", "Macbeth", "The Great Gatsby", "To Kill a Mockingbird",
            "Harry Potter", "Lord of the Rings", "Game of Thrones", "Star Wars", "Marvel Comics",
            "Academy Awards", "Oscars", "Grammy Awards", "Emmy Awards", "Golden Globe Awards",
            "Nobel Prize", "Pulitzer Prize", "Booker Prize", "Man Booker Prize"
        ],
        
        "Sports & Entertainment": [
            "Olympic Games", "Olympics", "Summer Olympics", "Winter Olympics", "FIFA World Cup",
            "Super Bowl", "World Series", "NBA Finals", "Stanley Cup", "Wimbledon",
            "Tour de France", "Formula One", "F1", "NASCAR", "Premier League",
            "Major League Baseball", "MLB", "National Football League", "NFL",
            "National Basketball Association", "NBA", "National Hockey League", "NHL"
        ],
        
        "Medical & Health": [
            "Human Immunodeficiency Virus", "HIV", "Acquired Immunodeficiency Syndrome", "AIDS",
            "Coronavirus Disease", "COVID-19", "Severe Acute Respiratory Syndrome", "SARS",
            "Centers for Disease Control", "CDC", "World Health Organization", "WHO",
            "Food and Drug Administration", "FDA", "American Medical Association", "AMA",
            "American Heart Association", "AHA", "American Cancer Society", "ACS"
        ],
        
        "Business & Economics": [
            "Gross Domestic Product", "GDP", "Gross National Product", "GNP",
            "Consumer Price Index", "CPI", "Unemployment Rate", "Inflation Rate",
            "Federal Reserve System", "Fed", "Wall Street", "Dow Jones", "S&P 500",
            "NASDAQ", "New York Stock Exchange", "NYSE", "London Stock Exchange",
            "Supply and Demand", "Market Economy", "Free Market", "Command Economy"
        ],
        
        "Time Periods & Eras": [
            "Stone Age", "Bronze Age", "Iron Age", "Middle Ages", "Dark Ages",
            "Renaissance Period", "Enlightenment Era", "Industrial Age", "Information Age",
            "Victorian Era", "Edwardian Era", "Roaring Twenties", "Great Depression Era",
            "Cold War Era", "Post-Cold War", "Modern Era", "Contemporary Period"
        ],
        
        "Educational Concepts": [
            "Bloom's Taxonomy", "Multiple Intelligences", "Learning Styles", "Constructivism",
            "Behaviorism", "Cognitivism", "Social Learning Theory", "Zone of Proximal Development",
            "Flipped Classroom", "Blended Learning", "Distance Learning", "Online Learning",
            "Massive Open Online Course", "MOOC", "Learning Management System", "LMS",
            "Student-Centered Learning", "Project-Based Learning", "Problem-Based Learning"
        ]
    }
    
    # Write data to Excel
    for category, nouns in categories.items():
        # Write category header
        ws.cell(row=row, column=1, value=category)
        ws.cell(row=row, column=1).fill = category_fill
        ws.cell(row=row, column=1).font = category_font
        ws.cell(row=row, column=1).alignment = Alignment(vertical="top")
        
        # Write proper nouns in this category
        for noun in nouns:
            ws.cell(row=row, column=2, value=noun)
            ws.cell(row=row, column=3, value="")  # Notes column for future use
            row += 1
        
        # Add blank row between categories
        row += 1
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save file
    output_path = Path(__file__).parent / "proper_nouns.xlsx"
    wb.save(output_path)
    print(f"✅ Generated {output_path} with {row - 2} proper noun entries across {len(categories)} categories")
    print(f"📊 Categories: {', '.join(categories.keys())}")
    
    return output_path

if __name__ == "__main__":
    generate_proper_nouns_excel()


