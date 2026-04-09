import os
import asyncio
import subprocess
from pathlib import Path
import cv2
import librosa
import numpy as np
from openai import OpenAI
from openai import APIError, APIConnectionError, APITimeoutError, RateLimitError
from typing import Dict, List, Any, Optional, Tuple
import base64
import json
import re
import logging
from dotenv import load_dotenv
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False



# Import enhanced configuration
from metrics_config import (
    ANALYSIS_CONFIG, FILLER_WORDS, SPEECH_METRICS, VISUAL_METRICS, PEDAGOGY_METRICS,
    calculate_metric_score, get_metric_feedback,
    MARS_CONFIG, MARS_RUBRIC_VERSION,
    compute_mars_content_category_score, compute_mars_content_category_score_detailed,
    compute_mars_delivery_category_score,
    compute_mars_engagement_category_score, compute_mars_overall_score,
)

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize OpenAI client
openai_client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))

class VideoAnalysisProcessor:
    def __init__(self):
        self.temp_dir = Path("temp_processing")
        self.temp_dir.mkdir(exist_ok=True)
        self.analysis_id = None  # Add this to store current analysis_id
        self.progress_callback = None  # Add this to store callback
    
    async def simulate_progress_during_operation(self, start_progress: int, end_progress: int, duration_seconds: float, message: str):
        """Simulate progress updates during a long operation"""
        if not self.progress_callback or not self.analysis_id:
            return
            
        steps = max(int(duration_seconds * 2), 1)  # Update every 0.5 seconds
        increment = (end_progress - start_progress) / steps
        
        for i in range(steps):
            # Check if stop was requested
            if hasattr(self, 'should_stop') and self.should_stop:
                raise Exception("Analysis stopped by user")
                
            current_progress = int(start_progress + (increment * i))
            await self.progress_callback(
                self.analysis_id, 
                current_progress, 
                f"{message} ({current_progress}%)"
            )
            await asyncio.sleep(0.3)  # Reduced sleep for more responsive updates
        
    async def process_video(self, video_path: Path, analysis_id: str, progress_callback, lecture_context: Optional[str] = None):
        """
        Enhanced processing pipeline for video analysis with improved sampling and metrics.
        lecture_context: optional user text (subject, topic, ILOs) passed to LLM calls for context-aware scoring.
        """
        try:
            # Store for use in helper methods
            self.analysis_id = analysis_id
            self.progress_callback = progress_callback
            self.lecture_context = (lecture_context or "").strip()[:20000] if lecture_context else ""

            logger.info(f"🎬 Starting enhanced video analysis for {analysis_id}")
            await progress_callback(analysis_id, 5, f"🎬 Starting enhanced video analysis for {analysis_id}")
            await asyncio.sleep(0)  # Force immediate execution
            
            logger.info(f"📁 File: {video_path.name} ({video_path.stat().st_size / (1024*1024):.1f}MB)")
            await progress_callback(analysis_id, 6, f"📁 File: {video_path.name} ({video_path.stat().st_size / (1024*1024):.1f}MB)")
            await asyncio.sleep(0)  # Force immediate execution
            
            # Step 1: Extract audio and video components with enhanced sampling
            logger.info("🔧 Step 1: Extracting audio and video components...")
            await progress_callback(analysis_id, 10, "🔧 Step 1: Extracting audio and video components...")
            await asyncio.sleep(0)  # Force immediate execution
            
            audio_path, video_frames = await self.extract_components(video_path)
            
            logger.info(f"✅ Extracted {len(video_frames)} video frames and audio track")
            await progress_callback(analysis_id, 20, f"✅ Extracted {len(video_frames)} video frames and audio track")
            await asyncio.sleep(0)  # Force immediate execution
            
            # Get video metadata
            cap_temp = cv2.VideoCapture(str(video_path))
            fps = cap_temp.get(cv2.CAP_PROP_FPS)
            total_frames = int(cap_temp.get(cv2.CAP_PROP_FRAME_COUNT))
            duration_seconds = total_frames / fps if fps > 0 else 0
            cap_temp.release()

            await progress_callback(analysis_id, 25, f"📹 Video duration: {duration_seconds:.1f}s, extracting {len(video_frames)} frames", {
                "step1": {
                    "status": "completed",
                    "duration": f"{duration_seconds:.1f}s",
                    "frames_extracted": len(video_frames),
                    "audio_format": "16kHz Mono WAV",
                    "video_resolution": "640x480"
                }
            })
            await asyncio.sleep(0)  # Force immediate execution
            
            # Step 2: Enhanced speech analysis with full transcript
            logger.info("🎤 Step 2: Analyzing speech with Whisper...")
            await progress_callback(analysis_id, 30, "🎤 Step 2: Analyzing speech with Whisper...")
            await asyncio.sleep(0)  # Force immediate execution
            
            speech_analysis = await self.analyze_speech_enhanced(audio_path)
            
            logger.info(f"✅ Speech analysis complete: {speech_analysis['word_count']} words, {speech_analysis['speaking_rate']:.1f} WPM")
            await progress_callback(analysis_id, 55, f"✅ Speech analysis complete: {speech_analysis['word_count']} words, {speech_analysis['speaking_rate']:.1f} WPM", {
                "step2": {
                    "status": "completed",
                    "word_count": speech_analysis['word_count'],
                    "transcript_length": len(speech_analysis['transcript']),
                    "speaking_rate": f"{speech_analysis['speaking_rate']:.1f} WPM",
                    "duration_minutes": f"{speech_analysis['duration_minutes']:.1f} min",
                    "filler_words": len(speech_analysis.get('filler_details', []))
                }
            })
            await asyncio.sleep(0)  # Force immediate execution
            
            # Step 3: Enhanced visual analysis with more frames
            logger.info("👁️ Step 3: Analyzing visual elements with GPT-4o Vision...")
            await progress_callback(analysis_id, 60, "👁️ Step 3: Analyzing visual elements with GPT-4o Vision...")
            
            visual_analysis = await self.analyze_visual_elements_enhanced(video_frames)
            
            logger.info(f"✅ Visual analysis complete: {visual_analysis.get('frames_analyzed', 0)} frames processed")
            await progress_callback(analysis_id, 75, f"✅ Visual analysis complete: {visual_analysis.get('frames_analyzed', 0)} frames processed", {
                "step3": {
                    "status": "completed",
                    "frames_analyzed": visual_analysis.get('frames_analyzed', 0),
                    "eye_contact": f"{visual_analysis.get('scores', {}).get('eye_contact', 0):.1f}/10",
                    "gestures": f"{visual_analysis.get('scores', {}).get('gestures', 0):.1f}/10",
                    "posture": f"{visual_analysis.get('scores', {}).get('posture', 0):.1f}/10"
                }
            })
            await asyncio.sleep(0)  # Force immediate execution
            
            # Step 4: Enhanced pedagogical assessment with full transcript
            logger.info("🎓 Step 4: Generating comprehensive pedagogical insights...")
            await progress_callback(analysis_id, 80, "🎓 Step 4: Generating comprehensive pedagogical insights...")
            await asyncio.sleep(0)  # Force immediate execution
            
            pedagogical_analysis = await self.analyze_pedagogy_enhanced(
                speech_analysis, visual_analysis, lecture_context=getattr(self, "lecture_context", "") or ""
            )
            
            logger.info("✅ Enhanced pedagogical analysis complete")
            await progress_callback(analysis_id, 90, "✅ Enhanced pedagogical analysis complete", {
                "step4": {
                    "status": "completed",
                    "content_organization": f"{pedagogical_analysis.get('content_organization', 0):.1f}/10",
                    "engagement": f"{pedagogical_analysis.get('engagement_techniques', 0):.1f}/10",
                    "teaching_effectiveness": f"{pedagogical_analysis.get('overall_effectiveness', 0):.1f}/10"
                }
            })
            await asyncio.sleep(0)  # Force immediate execution

            # Step 4.5: Analyze interaction and engagement
            logger.info("🤝 Step 4.5: Analyzing interaction and questioning techniques...")
            await progress_callback(analysis_id, 90, "🤝 Step 4.5: Analyzing interaction and questioning techniques...")
            
            # Student/audience question detection first (needed for SUI confidence)
            sf_metrics = await self.analyze_student_feedback_metrics(speech_analysis)
            speech_analysis["student_feedback_metrics"] = sf_metrics

            interaction_analysis = await self.analyze_interaction_engagement(speech_analysis)
            interaction_analysis.update(sf_metrics)

            logger.info(f"✅ Interaction analysis complete: {interaction_analysis['total_questions']} questions detected")
            await progress_callback(analysis_id, 92, f"✅ Interaction analysis complete: {interaction_analysis['total_questions']} questions detected")
            
            # Step 5: Extract sample frames for display (3-7 frames) at original resolution
            logger.info("📸 Step 5: Extracting sample frames for display...")
            sample_frames = []
            if video_frames:
                # Get video duration for even distribution
                cap = cv2.VideoCapture(str(video_path))
                fps = cap.get(cv2.CAP_PROP_FPS)
                total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
                duration_seconds = total_frames / fps if fps > 0 else 0
                
                # Select 3-7 evenly spaced frames across entire video duration
                num_frames_to_extract = min(7, max(3, len(video_frames)))
                
                # Calculate evenly spaced timestamps across entire video duration
                if duration_seconds > 0 and num_frames_to_extract > 0:
                    time_interval = duration_seconds / (num_frames_to_extract + 1)  # +1 to avoid first/last frame
                    selected_timestamps = [time_interval * (i + 1) for i in range(num_frames_to_extract)]
                else:
                    # Fallback to frame-based selection if duration unavailable
                    if len(video_frames) > num_frames_to_extract:
                        step = len(video_frames) // num_frames_to_extract
                        selected_indices = [i * step for i in range(num_frames_to_extract)]
                        if selected_indices[-1] >= len(video_frames):
                            selected_indices[-1] = len(video_frames) - 1
                        selected_timestamps = [video_frames[i].get('timestamp', 0) for i in selected_indices]
                    else:
                        selected_timestamps = [f.get('timestamp', 0) for f in video_frames]
                
                # Extract frames at calculated timestamps
                for target_timestamp in selected_timestamps:
                    # Find closest frame to target timestamp
                    closest_frame = min(video_frames, key=lambda f: abs(f.get('timestamp', 0) - target_timestamp))
                    frame_number = closest_frame.get('frame_number', 0)
                    timestamp = closest_frame.get('timestamp', 0)
                    
                    # Seek to the original frame in the video
                    cap.set(cv2.CAP_PROP_POS_FRAMES, frame_number)
                    ret, original_frame = cap.read()
                    
                    if ret and original_frame is not None:
                        # Use original frame (maintains aspect ratio)
                        frame = original_frame
                    else:
                        # Fallback to resized frame if original can't be read
                        frame = closest_frame.get('frame', None)
                        if frame is None:
                            continue  # Skip this frame if we can't get it
                    
                    # Convert frame to base64 (maintain original aspect ratio)
                    _, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 85])
                    frame_base64 = base64.b64encode(buffer).decode('utf-8')
                    
                    sample_frames.append({
                        'image': f"data:image/jpeg;base64,{frame_base64}",
                        'timestamp': round(timestamp, 2),
                        'frame_number': closest_frame.get('frame_number', frame_number)
                    })
                
                cap.release()
            
            logger.info(f"✅ Extracted {len(sample_frames)} sample frames at original resolution")
            
            # Step 6: Enhanced score combination with weighted sub-components
            logger.info("📊 Step 6: Calculating weighted component scores...")
            await progress_callback(analysis_id, 95, "📊 Step 6: Calculating weighted component scores...")
            
            final_results = await self.combine_analysis_enhanced(speech_analysis, visual_analysis, pedagogical_analysis, interaction_analysis, sample_frames)
            final_results["lecture_context"] = (getattr(self, "lecture_context", None) or "").strip()
            
            logger.info(f"✅ Enhanced analysis complete! Overall score: {final_results['overall_score']}/10")
            await progress_callback(analysis_id, 100, f"✅ Enhanced analysis complete! Overall score: {final_results['overall_score']}/10")
            await asyncio.sleep(0)  # Force immediate execution
            
            # Cleanup temporary files
            await self.cleanup_temp_files(audio_path, video_frames)
            logger.info("🧹 Cleanup complete")
            
            return final_results
            
        except Exception as e:
            logger.error(f"❌ Enhanced video analysis failed for {analysis_id}: {str(e)}")
            await progress_callback(analysis_id, 0, f"❌ Analysis failed: {str(e)}")
            await asyncio.sleep(0)  # Force immediate execution
            raise Exception(f"Enhanced video analysis failed: {str(e)}")
    
    async def extract_components(self, video_path: Path):
        """
        Enhanced component extraction with improved frame sampling
        """
        import shutil

        # Check if ffmpeg is available
        if not shutil.which('ffmpeg'):
            raise Exception("FFmpeg is not installed on the server. Please contact support.")
    
        # Extract audio using FFmpeg
        audio_path = self.temp_dir / f"{video_path.stem}_audio.wav"
        
        # Use subprocess to run FFmpeg
        audio_command = [
            'ffmpeg', '-i', str(video_path),
            '-acodec', 'pcm_s16le',
            '-ar', '16000',  # 16kHz sample rate for speech recognition
            '-ac', '1',      # Mono audio
            str(audio_path),
            '-y'  # Overwrite output files
        ]
        
        subprocess.run(audio_command, capture_output=True, check=True)
        
        # Enhanced video frame extraction with duration-based intervals
        video_frames = []
        cap = cv2.VideoCapture(str(video_path))
        fps = cap.get(cv2.CAP_PROP_FPS)
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        duration_seconds = total_frames / fps
        duration_minutes = duration_seconds / 60
        duration_hours = duration_minutes / 60
        
        # Frame sampling: for lectures ≥ 1 hour, use exactly 20 frames spread evenly across duration.
        # Otherwise: interval rules below, capped at 60 frames.
        if duration_seconds >= 3600:
            frames_to_extract = 20
            frame_interval_seconds = duration_seconds / frames_to_extract
        else:
            # 1. If video < 1 minute: extract every 20 seconds
            # 2. If video < 1 hour: extract every 1 minute (60 seconds)
            # 3. If video 1-2 hours (handled above): N/A
            if duration_minutes < 1.0:
                frame_interval_seconds = 20  # 20 seconds for videos < 1 minute
            elif duration_hours < 1.0:
                frame_interval_seconds = 60  # 1 minute
            else:
                frame_interval_seconds = 120  # 2 minutes (should not hit; ≥1h uses 20-frame rule)

            estimated_frames = int(duration_seconds / frame_interval_seconds)
            max_frames = 60
            frames_to_extract = min(estimated_frames, max_frames)
            if estimated_frames > max_frames:
                frame_interval_seconds = duration_seconds / max_frames
                frames_to_extract = max_frames
        
        frame_interval_frames = int(fps * frame_interval_seconds)
        
        logger.info(f"📊 Video duration: {duration_seconds:.1f}s ({duration_minutes:.1f} min), extracting {frames_to_extract} frames every {frame_interval_seconds:.0f} seconds")
        
        frame_count = 0
        extracted_count = 0
        next_extract_frame = 0
        
        while extracted_count < frames_to_extract:
            ret, frame = cap.read()
            if not ret:
                break
                
            # Extract frame at calculated intervals
            if frame_count >= next_extract_frame:
                # Resize frame for processing
                frame_resized = cv2.resize(frame, (640, 480))
                video_frames.append({
                    'frame': frame_resized,
                    'timestamp': frame_count / fps,
                    'frame_number': frame_count
                })
                extracted_count += 1
                next_extract_frame = frame_count + frame_interval_frames
                
            frame_count += 1
            
        cap.release()
        
        logger.info(f"📊 Extracted {len(video_frames)} frames from {duration_seconds:.1f}s video")
        return audio_path, video_frames
    
    def extract_timecoded_transcript(self, words_data: List[Dict]) -> List[Dict]:
        """
        Extract transcript with word-level timecodes
        Groups words by sentences - each sentence starts with a new timecode
        Uses punctuation marks (. ! ?) to identify sentence boundaries
        """
        timecoded_transcript = []
        current_sentence = []
        sentence_start_time = None
        
        # Ensure all word_data items are dictionaries (convert TranscriptionWord objects if needed)
        converted_words_data = []
        for word_item in words_data:
            if hasattr(word_item, 'word'):
                # It's a TranscriptionWord object - convert to dict
                converted_words_data.append({
                    'word': word_item.word,
                    'start': getattr(word_item, 'start', 0),
                    'end': getattr(word_item, 'end', 0)
                })
            elif isinstance(word_item, dict):
                converted_words_data.append(word_item)
            else:
                # Fallback
                converted_words_data.append({
                    'word': word_item.get('word', '') if isinstance(word_item, dict) else getattr(word_item, 'word', ''),
                    'start': word_item.get('start', 0) if isinstance(word_item, dict) else getattr(word_item, 'start', 0),
                    'end': word_item.get('end', 0) if isinstance(word_item, dict) else getattr(word_item, 'end', 0)
                })
        
        words_data = converted_words_data
        
        for word_data in words_data:
            word = word_data.get('word', '')
            word_start = word_data.get('start', 0)
            
            # If this is the start of a new sentence, set the sentence start time
            if sentence_start_time is None:
                sentence_start_time = word_start
            
            # Add word to current sentence
            current_sentence.append({
                'word': word,
                'start': word_start,
                'end': word_data.get('end', 0)
            })
            
            # Check if this word ends a sentence (period, exclamation, question mark)
            # Also check if word contains punctuation (Whisper might attach punctuation to words)
            sentence_ended = False
            if word:
                # Check if word ends with sentence-ending punctuation
                if word[-1] in ['.', '!', '?']:
                    sentence_ended = True
                # Also check if punctuation is attached (e.g., "word." or "word?")
                elif any(punct in word for punct in ['.', '!', '?']):
                    # Find the last punctuation mark
                    for punct in ['.', '!', '?']:
                        if punct in word:
                            sentence_ended = True
                            break
            
            # If sentence ended, finalize current sentence and start new one
            if sentence_ended and current_sentence:
                # Assign timestamp to all words in current sentence
                sentence_timestamp = self.format_timestamp(sentence_start_time)
                for sentence_word_data in current_sentence:
                    timecoded_transcript.append({
                        'word': sentence_word_data['word'],
                        'start': round(sentence_word_data['start'], 2),
                        'end': round(sentence_word_data['end'], 2),
                        'timestamp': sentence_timestamp
                    })
                
                # Start new sentence
                current_sentence = []
                sentence_start_time = None  # Will be set on next word
        
        # Add remaining words in final sentence (if any)
        if current_sentence:
            sentence_timestamp = self.format_timestamp(sentence_start_time) if sentence_start_time is not None else self.format_timestamp(0)
            for sentence_word_data in current_sentence:
                timecoded_transcript.append({
                    'word': sentence_word_data['word'],
                    'start': round(sentence_word_data['start'], 2),
                    'end': round(sentence_word_data['end'], 2),
                    'timestamp': sentence_timestamp
            })
        
        return timecoded_transcript
    
    def format_timestamp(self, seconds: float) -> str:
        """Format seconds to MM:SS or HH:MM:SS"""
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = int(seconds % 60)
        
        if hours > 0:
            return f"{hours:02d}:{minutes:02d}:{secs:02d}"
        else:
            return f"{minutes:02d}:{secs:02d}"
    
    async def transcribe_large_audio_file(self, audio_path: Path) -> Any:
        """
        Transcribe large audio files by splitting into chunks and processing separately
        Optimized for Railway server environment with memory constraints
        """
        import librosa
        import soundfile as sf
        import tempfile
        import os
        
        logger.info("🔧 Starting chunked audio processing...")
        
        # Load audio file
        audio_data, sample_rate = librosa.load(str(audio_path), sr=16000, mono=True)
        duration_seconds = len(audio_data) / sample_rate
        
        # Calculate optimal chunk parameters
        # Target: 20MB chunks (safe buffer under 25MB limit)
        # 16kHz mono WAV = 16,000 samples/sec * 2 bytes/sample = 32KB/sec
        # 20MB = 20 * 1024 * 1024 bytes = 20,971,520 bytes
        # 20MB / 32KB/sec = ~655 seconds = ~11 minutes
        target_chunk_duration = 600  # 10 minutes for safety
        overlap_duration = 30  # 30 seconds overlap
        
        chunk_samples = int(target_chunk_duration * sample_rate)
        overlap_samples = int(overlap_duration * sample_rate)
        
        logger.info(f"📊 Audio duration: {duration_seconds:.1f}s, chunking into ~{target_chunk_duration}s segments")
        
        # Process chunks
        all_transcripts = []
        all_words = []
        chunk_count = 0
        
        for start_sample in range(0, len(audio_data), chunk_samples - overlap_samples):
            chunk_count += 1
            end_sample = min(start_sample + chunk_samples, len(audio_data))
            
            # Extract chunk
            chunk_audio = audio_data[start_sample:end_sample]
            chunk_start_time = start_sample / sample_rate
            
            logger.info(f"📦 Processing chunk {chunk_count} ({chunk_start_time:.1f}s - {end_sample/sample_rate:.1f}s)")
            await self.progress_callback(
                self.analysis_id, 
                33 + int((chunk_count * 5) / max(1, (len(audio_data) // (chunk_samples - overlap_samples)))), 
                f"📦 Processing chunk {chunk_count}..."
            )
            
            # Save chunk to temporary file
            with tempfile.NamedTemporaryFile(suffix='.wav', delete=False) as temp_file:
                sf.write(temp_file.name, chunk_audio, sample_rate)
                temp_path = temp_file.name
            
            try:
                # Transcribe chunk (raw transcription, no punctuation prompt)
                with open(temp_path, "rb") as chunk_file:
                    chunk_response = openai_client.audio.transcriptions.create(
                        model="whisper-1",
                        file=chunk_file,
                        response_format="verbose_json",
                        timestamp_granularities=["word"],
                        language="en"  # Force English transcription (no punctuation prompt - get raw text)
                    )
                
                # Adjust timestamps to global time
                chunk_text = chunk_response.text
                chunk_words = getattr(chunk_response, 'words', [])
                
                # Convert TranscriptionWord objects to dictionaries and adjust timestamps
                converted_chunk_words = []
                for word_data in chunk_words:
                    if hasattr(word_data, 'word'):
                        # It's a TranscriptionWord object - convert to dict
                        converted_chunk_words.append({
                            'word': word_data.word,
                            'start': getattr(word_data, 'start', 0) + chunk_start_time,
                            'end': getattr(word_data, 'end', 0) + chunk_start_time
                        })
                    elif isinstance(word_data, dict):
                        # Already a dictionary - adjust timestamps
                        word_data['start'] += chunk_start_time
                        word_data['end'] += chunk_start_time
                        converted_chunk_words.append(word_data)
                    else:
                        # Fallback
                        start = word_data.get('start', 0) if isinstance(word_data, dict) else getattr(word_data, 'start', 0)
                        end = word_data.get('end', 0) if isinstance(word_data, dict) else getattr(word_data, 'end', 0)
                        converted_chunk_words.append({
                            'word': word_data.get('word', '') if isinstance(word_data, dict) else getattr(word_data, 'word', ''),
                            'start': start + chunk_start_time,
                            'end': end + chunk_start_time
                        })
                
                all_transcripts.append(chunk_text)
                all_words.extend(converted_chunk_words)
                
                logger.info(f"✅ Chunk {chunk_count} transcribed: {len(chunk_text)} chars")
                
            except Exception as e:
                logger.error(f"❌ Error processing chunk {chunk_count}: {e}")
                # Continue with other chunks
                continue
            finally:
                # Clean up temporary file
                try:
                    os.unlink(temp_path)
                except:
                    pass
        
        # Combine all transcripts
        combined_transcript = " ".join(all_transcripts)
        
        # Create a mock response object with combined data
        class CombinedResponse:
            def __init__(self, text, words):
                self.text = text
                self.words = words
        
        logger.info(f"✅ Chunked processing complete: {len(combined_transcript)} chars from {chunk_count} chunks")
        return CombinedResponse(combined_transcript, all_words)
    
    async def transcribe_large_audio_file_streaming(self, audio_path: Path) -> Any:
        """
        Memory-efficient streaming approach using FFmpeg to split audio
        Better for Railway's memory constraints with comprehensive error handling
        """
        import subprocess
        import tempfile
        import os
        import gc
        
        logger.info("🔧 Starting streaming chunked audio processing...")
        
        # Railway-specific optimizations
        max_chunks = 12  # Limit chunks for 1-hour max duration (10min chunks)
        max_retries = 3
        chunk_timeout = 300  # 5 minutes per chunk timeout
        
        # Get audio duration using FFprobe
        duration_cmd = [
            'ffprobe', '-v', 'quiet', '-show_entries', 'format=duration',
            '-of', 'csv=p=0', str(audio_path)
        ]
        
        try:
            duration_result = subprocess.run(duration_cmd, capture_output=True, text=True, check=True, timeout=30)
            duration_seconds = float(duration_result.stdout.strip())
            
            # Validate duration for Railway constraints
            if duration_seconds > 3600:  # 1 hour limit
                raise ValueError(f"Audio duration ({duration_seconds:.1f}s) exceeds 1-hour limit")
                
        except subprocess.TimeoutExpired:
            logger.error("❌ FFprobe timeout - audio file may be corrupted")
            raise Exception("Audio file analysis timeout")
        except Exception as e:
            logger.error(f"❌ Could not get audio duration: {e}")
            # Fallback to librosa method
            return await self.transcribe_large_audio_file(audio_path)
        
        # Calculate chunk parameters
        target_chunk_duration = 600  # 10 minutes
        overlap_duration = 30  # 30 seconds
        
        all_transcripts = []
        all_words = []
        chunk_count = 0
        
        # Calculate total expected chunks
        total_expected_chunks = min(max_chunks, int(duration_seconds / (target_chunk_duration - overlap_duration)) + 1)
        
        for start_time in range(0, int(duration_seconds), target_chunk_duration - overlap_duration):
            if chunk_count >= max_chunks:
                logger.warning(f"⚠️ Reached maximum chunk limit ({max_chunks}), stopping processing")
                break
                
            chunk_count += 1
            end_time = min(start_time + target_chunk_duration, duration_seconds)
            
            logger.info(f"📦 Processing chunk {chunk_count}/{total_expected_chunks} ({start_time}s - {end_time}s)")
            await self.progress_callback(
                self.analysis_id, 
                33 + int((chunk_count * 5) / total_expected_chunks), 
                f"📦 Processing chunk {chunk_count}/{total_expected_chunks}..."
            )
            
            # Create temporary file for chunk
            temp_path = None
            retry_count = 0
            
            while retry_count < max_retries:
                try:
                    with tempfile.NamedTemporaryFile(suffix='.wav', delete=False) as temp_file:
                        temp_path = temp_file.name
                    
                    # Extract chunk using FFmpeg with timeout
                    chunk_cmd = [
                        'ffmpeg', '-i', str(audio_path),
                        '-ss', str(start_time),
                        '-t', str(end_time - start_time),
                        '-acodec', 'pcm_s16le',
                        '-ar', '16000',
                        '-ac', '1',
                        temp_path,
                        '-y'
                    ]
                    
                    subprocess.run(chunk_cmd, capture_output=True, check=True, timeout=chunk_timeout)
                    
                    # Check chunk file size
                    chunk_size_mb = os.path.getsize(temp_path) / (1024 * 1024)
                    if chunk_size_mb > 25:
                        logger.warning(f"⚠️ Chunk {chunk_count} is {chunk_size_mb:.1f}MB, may exceed API limit")
                    
                    # Transcribe chunk with retry logic (raw transcription, no punctuation prompt)
                    chunk_success = False
                    for api_retry in range(max_retries):
                        try:
                            with open(temp_path, "rb") as chunk_file:
                                chunk_response = openai_client.audio.transcriptions.create(
                                    model="whisper-1",
                                    file=chunk_file,
                                    response_format="verbose_json",
                                    timestamp_granularities=["word"],
                                    language="en"  # Force English transcription (no punctuation prompt - get raw text)
                                )
                            
                            # Adjust timestamps to global time
                            chunk_text = chunk_response.text
                            chunk_words = getattr(chunk_response, 'words', [])
                            
                            # Convert TranscriptionWord objects to dictionaries and adjust timestamps
                            converted_chunk_words = []
                            for word_data in chunk_words:
                                if hasattr(word_data, 'word'):
                                    # It's a TranscriptionWord object - convert to dict
                                    converted_chunk_words.append({
                                        'word': word_data.word,
                                        'start': getattr(word_data, 'start', 0) + start_time,
                                        'end': getattr(word_data, 'end', 0) + start_time
                                    })
                                elif isinstance(word_data, dict):
                                    # Already a dictionary - adjust timestamps
                                    word_data['start'] += start_time
                                    word_data['end'] += start_time
                                    converted_chunk_words.append(word_data)
                                else:
                                    # Fallback
                                    start = word_data.get('start', 0) if isinstance(word_data, dict) else getattr(word_data, 'start', 0)
                                    end = word_data.get('end', 0) if isinstance(word_data, dict) else getattr(word_data, 'end', 0)
                                    converted_chunk_words.append({
                                        'word': word_data.get('word', '') if isinstance(word_data, dict) else getattr(word_data, 'word', ''),
                                        'start': start + start_time,
                                        'end': end + start_time
                                    })
                            
                            all_transcripts.append(chunk_text)
                            all_words.extend(converted_chunk_words)
                            
                            logger.info(f"✅ Chunk {chunk_count} transcribed: {len(chunk_text)} chars")
                            chunk_success = True
                            break
                            
                        except Exception as api_error:
                            logger.warning(f"⚠️ API error for chunk {chunk_count}, retry {api_retry + 1}: {api_error}")
                            if api_retry < max_retries - 1:
                                await asyncio.sleep(2 ** api_retry)  # Exponential backoff
                    
                    if chunk_success:
                        break
                    else:
                        raise Exception("All API retries failed")
                        
                except subprocess.TimeoutExpired:
                    logger.error(f"❌ FFmpeg timeout for chunk {chunk_count}")
                    retry_count += 1
                    if retry_count < max_retries:
                        await asyncio.sleep(2)
                        continue
                except subprocess.CalledProcessError as e:
                    logger.error(f"❌ FFmpeg error processing chunk {chunk_count}: {e}")
                    retry_count += 1
                    if retry_count < max_retries:
                        await asyncio.sleep(2)
                        continue
                except Exception as e:
                    logger.error(f"❌ Error processing chunk {chunk_count}: {e}")
                    retry_count += 1
                    if retry_count < max_retries:
                        await asyncio.sleep(2)
                        continue
                finally:
                    # Clean up temporary file
                    if temp_path and os.path.exists(temp_path):
                        try:
                            os.unlink(temp_path)
                        except:
                            pass
                    
                    # Force garbage collection for Railway memory management
                    gc.collect()
            
            if retry_count >= max_retries:
                logger.error(f"❌ Failed to process chunk {chunk_count} after {max_retries} retries")
                continue
        
        # Combine all transcripts
        combined_transcript = " ".join(all_transcripts)
        
        # Create a mock response object with combined data
        class CombinedResponse:
            def __init__(self, text, words):
                self.text = text
                self.words = words
        
        logger.info(f"✅ Streaming chunked processing complete: {len(combined_transcript)} chars from {chunk_count} chunks")
        return CombinedResponse(combined_transcript, all_words)
    
    async def analyze_speech_enhanced(self, audio_path: Path) -> Dict[str, Any]:
        """Enhanced speech analysis using chunked Whisper transcription for large files"""
        logger.info("🎤 Starting enhanced Whisper transcription...")
        await self.progress_callback(self.analysis_id, 30, "🎤 Starting enhanced Whisper transcription (English)...")
        
        # Check file size and determine if chunking is needed
        file_size_mb = audio_path.stat().st_size / (1024 * 1024)
        logger.info(f"📊 Audio file size: {file_size_mb:.1f}MB")
        
        if file_size_mb > 20:  # Use 20MB threshold for safety
            logger.info("📦 Large audio file detected, using chunked processing...")
            await self.progress_callback(self.analysis_id, 32, f"📦 Large audio file ({file_size_mb:.1f}MB), processing in chunks...")
            
            # Use streaming approach for better memory efficiency on Railway
            try:
                transcript_response = await self.transcribe_large_audio_file_streaming(audio_path)
            except Exception as e:
                logger.warning(f"⚠️ Streaming approach failed: {e}, falling back to librosa method")
                transcript_response = await self.transcribe_large_audio_file(audio_path)
        else:
            logger.info("📄 Small audio file, processing directly...")
            await self.progress_callback(self.analysis_id, 35, "📄 Processing audio file directly (English)...")
            
            # Process directly with English language enforcement (raw transcription, no punctuation prompt)
            with open(audio_path, "rb") as audio_file:
                transcript_response = openai_client.audio.transcriptions.create(
                    model="whisper-1",
                    file=audio_file,
                    response_format="verbose_json",
                    timestamp_granularities=["word"],
                    language="en"  # Force English transcription (no punctuation prompt - get raw text)
                )
        
        await self.progress_callback(self.analysis_id, 40, "✅ Whisper transcription complete")
        
        logger.info("✅ Whisper transcription complete")
        transcript_text = transcript_response.text
        logger.info(f"📝 Full transcript length: {len(transcript_text)} characters")
        
        # Log first 200 characters for debugging
        preview = transcript_text[:200] + "..." if len(transcript_text) > 200 else transcript_text
        logger.info(f"📄 Transcript preview: {preview}")
        
        await self.progress_callback(self.analysis_id, 42, f"📝 Full transcript length: {len(transcript_text)} characters")
        
        # Post-process transcript with GPT-5-nano for punctuation and sentence segmentation
        logger.info("📝 Post-processing transcript with GPT-5-nano for punctuation and sentence segmentation...")
        await self.progress_callback(self.analysis_id, 42.5, "📝 Post-processing transcript for proper punctuation and sentence segmentation...")
        transcript_text = await self.post_process_transcript_with_gpt(transcript_text)
        logger.info("✅ Transcript post-processing complete")
        
        # Get word-level data early (needed for intonation analysis)
        # Handle both single response and chunked responses (CombinedResponse)
        try:
            if hasattr(transcript_response, 'words'):
                words_data = transcript_response.words
            elif isinstance(transcript_response, dict) and 'words' in transcript_response:
                words_data = transcript_response['words']
            else:
                # Try to get words from the response object
                words_data = getattr(transcript_response, 'words', [])
            
            # Ensure words_data is a list
            if words_data is None:
                words_data = []
            
            # Convert TranscriptionWord objects to dictionaries if needed
            # Whisper API returns TranscriptionWord objects with attributes, not dicts
            converted_words_data = []
            for word_item in words_data:
                if hasattr(word_item, 'word'):
                    # It's a TranscriptionWord object - convert to dict
                    converted_words_data.append({
                        'word': word_item.word,
                        'start': getattr(word_item, 'start', 0),
                        'end': getattr(word_item, 'end', 0)
                    })
                elif isinstance(word_item, dict):
                    # Already a dictionary
                    converted_words_data.append(word_item)
                else:
                    # Fallback: try to access as dict or object
                    converted_words_data.append({
                        'word': word_item.get('word', '') if isinstance(word_item, dict) else getattr(word_item, 'word', ''),
                        'start': word_item.get('start', 0) if isinstance(word_item, dict) else getattr(word_item, 'start', 0),
                        'end': word_item.get('end', 0) if isinstance(word_item, dict) else getattr(word_item, 'end', 0)
                    })
            
            words_data = converted_words_data
            logger.info(f"📊 Word timestamps available: {len(words_data)} words")
        except Exception as e:
            logger.warning(f"⚠️ Error extracting word timestamps: {e}")
            words_data = []
        
        # Enhanced speech metrics calculation
        logger.info("🔢 Calculating enhanced speech metrics...")
        await self.progress_callback(self.analysis_id, 43, "🔢 Calculating enhanced speech metrics...")
        
        audio_data, sample_rate = librosa.load(str(audio_path), sr=16000)
        
        # Basic metrics
        duration_minutes = len(audio_data) / sample_rate / 60
        words = transcript_text.split()
        word_count = len(words)
        speaking_rate = word_count / duration_minutes if duration_minutes > 0 else 0
        
        logger.info(f"📊 Speaking rate: {speaking_rate:.1f} WPM")
        await self.progress_callback(self.analysis_id, 45, f"📊 Speaking rate: {speaking_rate:.1f} WPM")
        
        # Enhanced voice activity detection
        voice_activity = librosa.effects.split(audio_data, top_db=20)
        speaking_time = sum([(end - start) / sample_rate for start, end in voice_activity])
        speaking_ratio = speaking_time / (len(audio_data) / sample_rate)
        
        # Note: Intonation analysis removed - GPT post-processing handles punctuation better
        # Enhanced filler word analysis with timecodes
        text_lower = transcript_text.lower()
        filler_count = 0
        filler_details = []
        filler_timecodes = []
        
        # Ensure words_data contains dictionaries (convert TranscriptionWord objects if needed)
        converted_words_for_filler = []
        for word_item in words_data:
            if hasattr(word_item, 'word'):
                converted_words_for_filler.append({
                    'word': word_item.word,
                    'start': getattr(word_item, 'start', 0),
                    'end': getattr(word_item, 'end', 0)
                })
            elif isinstance(word_item, dict):
                converted_words_for_filler.append(word_item)
            else:
                converted_words_for_filler.append({
                    'word': word_item.get('word', '') if isinstance(word_item, dict) else getattr(word_item, 'word', ''),
                    'start': word_item.get('start', 0) if isinstance(word_item, dict) else getattr(word_item, 'start', 0),
                    'end': word_item.get('end', 0) if isinstance(word_item, dict) else getattr(word_item, 'end', 0)
                })
        
        for filler in FILLER_WORDS:
            count = 0
            for word_data in converted_words_for_filler:
                word = word_data.get('word', '').lower().strip('.,!?')
                if word == filler:
                    count += 1
                    filler_timecodes.append({
                        'word': filler,
                        'timestamp': self.format_timestamp(word_data.get('start', 0)),
                        'start': round(word_data.get('start', 0), 2),
                        'end': round(word_data.get('end', 0), 2)
                    })
            
            if count > 0:
                filler_count += count
                filler_details.append({'word': filler, 'count': count})
        
        filler_ratio = filler_count / word_count if word_count > 0 else 0
        
        logger.info(f"📊 Enhanced filler analysis: {filler_ratio:.3f} ({filler_count} filler words)")
        await self.progress_callback(self.analysis_id, 47, f"📊 Enhanced filler analysis: {filler_ratio:.3f} ({filler_count} filler words)")
        
        # Voice variety analysis (pitch and energy variation)
        voice_variety_score = self.calculate_voice_variety(audio_data, sample_rate)
        
        # Pause effectiveness analysis
        pause_effectiveness_score = self.calculate_pause_effectiveness(transcript_response.words if hasattr(transcript_response, 'words') else [])
        
        # Extract key phrases using improved analysis
        sentences = re.split(r'[.!?]+', transcript_text)
        highlights = [s.strip() for s in sentences if len(s.strip()) > 50][:10]
        
        logger.info("🎓 Analyzing full transcript structure with GPT-4o...")
        await self.progress_callback(self.analysis_id, 48, "🎓 Analyzing full transcript structure with GPT-4o...")
        
        content_analysis = await self.analyze_content_structure_enhanced(transcript_text)
        
        logger.info("✅ Enhanced content structure analysis complete")
        await self.progress_callback(self.analysis_id, 50, "✅ Enhanced content structure analysis complete")
        
        return {
            'transcript': transcript_text,
            'timecoded_transcript': self.extract_timecoded_transcript(words_data),
            'filler_timecodes': filler_timecodes,
            'confidence': 0.95,
            'speaking_rate': speaking_rate,
            'speaking_ratio': speaking_ratio,
            'filler_ratio': filler_ratio,
            'filler_details': filler_details,
            'voice_variety_score': voice_variety_score,
            'pause_effectiveness_score': pause_effectiveness_score,
            'word_count': word_count,
            'duration_minutes': duration_minutes,
            'highlights': highlights,
            'content_structure': content_analysis,
            'word_timestamps': words_data  # Already converted to dictionaries above
        }
    
    def calculate_voice_variety(self, audio_data: np.ndarray, sample_rate: int) -> float:
        """
        Calculate voice variety based on pitch and energy variations
        """
        try:
            # Extract pitch (fundamental frequency)
            pitches, magnitudes = librosa.piptrack(y=audio_data, sr=sample_rate, threshold=0.1)
            
            # Get the most prominent pitch at each time frame
            pitch_values = []
            for t in range(pitches.shape[1]):
                index = magnitudes[:, t].argmax()
                pitch = pitches[index, t]
                if pitch > 0:  # Valid pitch detected
                    pitch_values.append(pitch)
            
            if len(pitch_values) < 10:  # Not enough data
                return 0.5
            
            # Calculate pitch variation coefficient
            pitch_std = np.std(pitch_values)
            pitch_mean = np.mean(pitch_values)
            pitch_variation = pitch_std / pitch_mean if pitch_mean > 0 else 0
            
            # Calculate energy variation
            energy = librosa.feature.rms(y=audio_data)[0]
            energy_variation = np.std(energy) / np.mean(energy) if np.mean(energy) > 0 else 0
            
            # Combine pitch and energy variations (normalized to 0-1 scale)
            voice_variety = min(1.0, (pitch_variation * 2 + energy_variation) / 2)
            
            return voice_variety
            
        except Exception as e:
            logger.warning(f"Voice variety calculation failed: {e}")
            return 0.5  # Default moderate score
    
    def calculate_pause_effectiveness(self, word_timestamps: List[Dict]) -> float:
        """
        Calculate pause effectiveness based on timing and context
        """
        if len(word_timestamps) < 5:
            return 0.5
        
        try:
            strategic_pauses = 0
            excessive_pauses = 0
            total_pauses = 0
            
            for i in range(len(word_timestamps) - 1):
                current_word = word_timestamps[i]
                next_word = word_timestamps[i + 1]
                
                current_end = current_word.get('end', 0)
                next_start = next_word.get('start', 0)
                pause_duration = next_start - current_end
                
                if pause_duration > 0.5:  # Pause longer than 0.5 seconds
                    total_pauses += 1
                    
                    if 0.5 <= pause_duration <= 2.0:  # Strategic pause range
                        strategic_pauses += 1
                    elif pause_duration > 3.0:  # Excessive pause
                        excessive_pauses += 1
            
            if total_pauses == 0:
                return 0.3  # No pauses might indicate rushed delivery
            
            # Calculate effectiveness ratio
            effectiveness = (strategic_pauses - excessive_pauses * 0.5) / total_pauses
            return max(0.0, min(1.0, effectiveness))
            
        except Exception as e:
            logger.warning(f"Pause effectiveness calculation failed: {e}")
            return 0.5
    
    async def post_process_transcript_with_gpt(self, raw_transcript: str) -> str:
        """
        Post-process raw Whisper transcript with GPT-5-nano to:
        - Add proper punctuation (commas, periods, question marks)
        - Segment into natural sentences
        - Preserve original meaning
        - Do not rewrite or improve wording
        """
        try:
            # Split transcript into chunks if too long (GPT-5-nano has token limits)
            max_chunk_length = 8000  # Conservative limit to leave room for prompt and response
            chunks = []
            
            if len(raw_transcript) <= max_chunk_length:
                chunks = [raw_transcript]
            else:
                # Split by approximate word count (roughly 5 chars per word)
                words = raw_transcript.split()
                words_per_chunk = max_chunk_length // 5
                
                for i in range(0, len(words), words_per_chunk):
                    chunk = ' '.join(words[i:i + words_per_chunk])
                    chunks.append(chunk)
            
            processed_chunks = []
            
            for idx, chunk in enumerate(chunks):
                if len(chunks) > 1:
                    logger.info(f"📝 Processing transcript chunk {idx + 1}/{len(chunks)}...")
                
                response = openai_client.chat.completions.create(
                    model="gpt-5-nano",
                    messages=[
                        {
                            "role": "system",
                            "content": """You are a transcript post-processing assistant. Your task is to add proper punctuation and segment sentences in a raw transcript.

IMPORTANT RULES:
1. Add proper punctuation: commas, periods, question marks, exclamation marks
2. Segment into natural sentences (each sentence should start with a capital letter)
3. Preserve the original meaning exactly - do not rewrite or improve wording
4. Do not add or remove words
5. Do not change the order of words
6. Only add punctuation and capitalization
7. Use question marks (?) only for actual questions (not for words like "what" or "why" in the middle of statements)
8. Maintain the original vocabulary and phrasing"""
                        },
                        {
                            "role": "user",
                            "content": f"""Please add proper punctuation and segment this raw transcript into natural sentences. Preserve the exact wording and meaning - only add punctuation and capitalization.

Raw transcript:
{chunk}

Return only the processed transcript with proper punctuation and sentence segmentation. Do not include any explanations or additional text."""
                        }
                    ],
                    max_completion_tokens=4000
                    # Note: GPT-5-nano only supports default temperature (1), cannot set custom values
                )
                
                if response.choices and response.choices[0].message.content:
                    processed_chunk = response.choices[0].message.content.strip()
                    processed_chunks.append(processed_chunk)
                else:
                    logger.warning(f"⚠️ GPT post-processing returned empty response for chunk {idx + 1}, using original")
                    processed_chunks.append(chunk)
            
            # Combine processed chunks
            processed_transcript = ' '.join(processed_chunks)
            
            logger.info(f"✅ Post-processed transcript: {len(processed_transcript)} characters (original: {len(raw_transcript)} characters)")
            return processed_transcript
            
        except Exception as e:
            logger.warning(f"⚠️ GPT post-processing failed: {e}. Using original transcript.")
            return raw_transcript
    
    def analyze_intonation_for_questions(self, audio_data: np.ndarray, sample_rate: int, word_timestamps: List[Dict]) -> Dict[str, Any]:
        """
        Analyze intonation patterns to detect questions based on rising pitch at end of sentences.
        Returns timestamps where rising intonation suggests a question.
        """
        try:
            # Handle empty or None word_timestamps
            if not word_timestamps or len(word_timestamps) == 0:
                logger.warning("⚠️ No word timestamps available for intonation analysis")
                return {'question_timestamps': []}
            # Extract pitch (fundamental frequency) using piptrack
            pitches, magnitudes = librosa.piptrack(y=audio_data, sr=sample_rate, threshold=0.1)
            
            # Get time frames for pitch analysis
            times = librosa.frames_to_time(np.arange(pitches.shape[1]), sr=sample_rate)
            
            # Extract pitch values over time
            pitch_track = []
            for t in range(pitches.shape[1]):
                index = magnitudes[:, t].argmax()
                pitch = pitches[index, t]
                if pitch > 0:  # Valid pitch detected
                    pitch_track.append((times[t], pitch))
                else:
                    pitch_track.append((times[t], None))
            
            if len(pitch_track) < 10:
                return {'question_timestamps': []}
            
            # Convert word_timestamps to dictionaries if needed (do this once at the start)
            converted_word_timestamps_for_pauses = []
            if word_timestamps:
                for word_item in word_timestamps:
                    if hasattr(word_item, 'word'):
                        converted_word_timestamps_for_pauses.append({
                            'word': word_item.word,
                            'start': getattr(word_item, 'start', 0),
                            'end': getattr(word_item, 'end', 0)
                        })
                    elif isinstance(word_item, dict):
                        converted_word_timestamps_for_pauses.append(word_item)
                    else:
                        converted_word_timestamps_for_pauses.append({
                            'word': word_item.get('word', '') if isinstance(word_item, dict) else getattr(word_item, 'word', ''),
                            'start': word_item.get('start', 0) if isinstance(word_item, dict) else getattr(word_item, 'start', 0),
                            'end': word_item.get('end', 0) if isinstance(word_item, dict) else getattr(word_item, 'end', 0)
                        })
                word_timestamps = converted_word_timestamps_for_pauses
            
            # Find sentence boundaries from word timestamps
            sentence_end_timestamps = []
            if word_timestamps:
                # Group words into sentences (look for pauses > 0.8 seconds or punctuation)
                current_sentence_end = None
                for i in range(len(word_timestamps) - 1):
                    current_word = word_timestamps[i]
                    next_word = word_timestamps[i + 1]
                    
                    word_text = current_word.get('word', '').strip()
                    pause_duration = next_word.get('start', 0) - current_word.get('end', 0)
                    
                    # Check if this word ends with punctuation (from Whisper)
                    if any(punct in word_text for punct in ['.', '!', '?']):
                        current_sentence_end = current_word.get('end', 0)
                        sentence_end_timestamps.append(current_sentence_end)
                    elif pause_duration > 0.8:  # Long pause suggests sentence end
                        current_sentence_end = current_word.get('end', 0)
                        sentence_end_timestamps.append(current_sentence_end)
                
                # Ensure word_timestamps contains dictionaries before accessing
                converted_word_timestamps_for_sentences = []
                for word_item in word_timestamps:
                    if hasattr(word_item, 'word'):
                        converted_word_timestamps_for_sentences.append({
                            'word': word_item.word,
                            'start': getattr(word_item, 'start', 0),
                            'end': getattr(word_item, 'end', 0)
                        })
                    elif isinstance(word_item, dict):
                        converted_word_timestamps_for_sentences.append(word_item)
                    else:
                        converted_word_timestamps_for_sentences.append({
                            'word': word_item.get('word', '') if isinstance(word_item, dict) else getattr(word_item, 'word', ''),
                            'start': word_item.get('start', 0) if isinstance(word_item, dict) else getattr(word_item, 'start', 0),
                            'end': word_item.get('end', 0) if isinstance(word_item, dict) else getattr(word_item, 'end', 0)
                        })
                
                # Add final word as sentence end
                if converted_word_timestamps_for_sentences:
                    sentence_end_timestamps.append(converted_word_timestamps_for_sentences[-1].get('end', 0))
            
            # Analyze pitch at sentence endings for rising intonation
            question_timestamps = []
            window_size = 0.5  # Analyze last 0.5 seconds of each sentence
            
            for sentence_end_time in sentence_end_timestamps:
                # Get pitch values in the window before sentence end
                window_start = max(0, sentence_end_time - window_size)
                window_end = sentence_end_time
                
                # Extract pitch values in this window
                window_pitches = []
                for time, pitch in pitch_track:
                    if window_start <= time <= window_end and pitch is not None:
                        window_pitches.append(pitch)
                
                if len(window_pitches) < 3:
                    continue
                
                # Check for rising intonation (pitch increases toward end)
                # Split window into first half and second half
                mid_point = len(window_pitches) // 2
                first_half_pitch = np.mean(window_pitches[:mid_point]) if mid_point > 0 else 0
                second_half_pitch = np.mean(window_pitches[mid_point:]) if len(window_pitches) > mid_point else 0
                
                # Rising intonation: second half pitch is significantly higher
                pitch_rise_ratio = (second_half_pitch - first_half_pitch) / first_half_pitch if first_half_pitch > 0 else 0
                
                # Threshold: 15% pitch rise suggests a question
                if pitch_rise_ratio > 0.15:
                    question_timestamps.append({
                        'timestamp': sentence_end_time,
                        'pitch_rise': pitch_rise_ratio,
                        'confidence': min(1.0, pitch_rise_ratio / 0.3)  # Normalize to 0-1
                    })
            
            return {
                'question_timestamps': question_timestamps,
                'total_sentences': len(sentence_end_timestamps),
                'questions_detected': len(question_timestamps)
            }
            
        except Exception as e:
            logger.warning(f"Intonation analysis failed: {e}")
            return {'question_timestamps': []}
    
    def enhance_transcript_with_intonation(self, transcript: str, intonation_data: Dict[str, Any], word_timestamps: List[Dict]) -> str:
        """
        Enhance transcript by adding question marks based on intonation analysis.
        Only adds question marks if they're not already present.
        """
        try:
            if not intonation_data or not intonation_data.get('question_timestamps'):
                return transcript
            
            question_timestamps = intonation_data['question_timestamps']
            
            # Create a mapping of timestamps to word positions
            # Find words near question timestamps and add question marks
            enhanced_transcript = transcript
            changes_made = 0
            
            # Sort question timestamps by confidence (highest first)
            sorted_questions = sorted(question_timestamps, key=lambda x: x.get('confidence', 0), reverse=True)
            
            # For each detected question, find the corresponding word and add question mark
            for question_info in sorted_questions:
                question_time = question_info['timestamp']
                confidence = question_info.get('confidence', 0)
                
                # Only process high-confidence questions (confidence > 0.5)
                if confidence < 0.5:
                    continue
                
                # Find the word closest to this timestamp
                closest_word_idx = None
                min_time_diff = float('inf')
                
                # Ensure word_timestamps contains dictionaries
                converted_word_timestamps = []
                for word_item in word_timestamps:
                    if hasattr(word_item, 'word'):
                        converted_word_timestamps.append({
                            'word': word_item.word,
                            'start': getattr(word_item, 'start', 0),
                            'end': getattr(word_item, 'end', 0)
                        })
                    elif isinstance(word_item, dict):
                        converted_word_timestamps.append(word_item)
                    else:
                        converted_word_timestamps.append({
                            'word': word_item.get('word', '') if isinstance(word_item, dict) else getattr(word_item, 'word', ''),
                            'start': word_item.get('start', 0) if isinstance(word_item, dict) else getattr(word_item, 'start', 0),
                            'end': word_item.get('end', 0) if isinstance(word_item, dict) else getattr(word_item, 'end', 0)
                        })
                
                for idx, word_data in enumerate(converted_word_timestamps):
                    word_end = word_data.get('end', 0)
                    time_diff = abs(word_end - question_time)
                    if time_diff < min_time_diff:
                        min_time_diff = time_diff
                        closest_word_idx = idx
                
                if closest_word_idx is not None and min_time_diff < 1.0:  # Within 1 second
                    word_data = converted_word_timestamps[closest_word_idx]
                    word_text = word_data.get('word', '').strip()
                    
                    # Check if word already has punctuation
                    if '?' not in word_text and '.' not in word_text and '!' not in word_text:
                        # Find this word in the transcript and add question mark
                        # Use regex to find the word at the end of a sentence
                        word_pattern = re.escape(word_text)
                        # Look for the word followed by space or end of string
                        pattern = rf'\b{word_pattern}\b(?=\s|$)'
                        
                        # Replace only if not already a question
                        if re.search(pattern, enhanced_transcript):
                            # Replace the word with word + question mark
                            enhanced_transcript = re.sub(
                                pattern,
                                lambda m: word_text + '?',
                                enhanced_transcript,
                                count=1  # Only replace first occurrence
                            )
                            changes_made += 1
            
            if changes_made > 0:
                logger.info(f"✅ Enhanced transcript with {changes_made} question marks based on intonation")
            
            return enhanced_transcript
            
        except Exception as e:
            logger.warning(f"Transcript enhancement with intonation failed: {e}")
            return transcript
    
    async def analyze_content_structure_enhanced(self, transcript: str) -> Dict[str, Any]:
        """
        Enhanced content structure analysis using FULL transcript
        """
        # Use full transcript instead of limiting to 3000 characters
        full_transcript = transcript
        logger.info(f"📊 Analyzing full transcript: {len(full_transcript)} characters")
        
        response = openai_client.chat.completions.create(
            model="gpt-5-nano",
            messages=[
                {
                    "role": "system",
                    "content": """You are an expert in educational content analysis. Analyze the COMPLETE lecture transcript for:
                    1. Content organization and logical flow throughout the entire lecture
                    2. Use of examples and explanations across all content
                    3. Educational techniques used throughout
                    4. Clarity of key concepts across the full lecture
                    5. Student engagement elements throughout the session
                    6. Knowledge checking and comprehension verification
                    7. Conclusion and summary effectiveness
                    
                    Provide detailed scores (1-10) and specific feedback based on the COMPLETE transcript."""
                },
                {
                    "role": "user",
                    "content": f"""Analyze this COMPLETE lecture transcript for educational quality:
                    
                    {full_transcript}
                    
                    Rate the following aspects (1-10) based on the ENTIRE lecture:
                    - Content organization (logical flow, transitions, structure)
                    - Engagement techniques (questions, interactions, variety)
                    - Communication clarity (explanations, terminology, pace)
                    - Use of examples (quantity, quality, relevance)
                    - Knowledge checking (comprehension verification, feedback)
                    
                    Also identify:
                    - Key topics covered throughout
                    - Teaching techniques used across the session
                    - Specific areas for improvement
                    - Strengths demonstrated throughout
                    
                    Return as JSON with keys: content_organization, engagement_techniques, communication_clarity, use_of_examples, knowledge_checking, key_topics, techniques, improvements, strengths"""
                }
            ],
            max_completion_tokens=1200  # Increased token limit for full analysis
        )
        
        try:
            # Check if response content exists
            response_content = response.choices[0].message.content
            if not response_content:
                raise ValueError("AI response content is None or empty")
            
            return self._safe_json_loads(response_content)
        except (json.JSONDecodeError, ValueError, AttributeError):
            return {
                'content_organization': 7.5,
                'engagement_techniques': 7.0,
                'communication_clarity': 7.5,
                'use_of_examples': 7.0,
                'knowledge_checking': 6.5,
                'key_topics': ['Main lecture content analyzed'],
                'techniques': ['Direct explanation', 'Content delivery'],
                'improvements': ['More examples needed', 'Improve engagement', 'Add comprehension checks'],
                'strengths': ['Clear delivery', 'Good organization', 'Professional presentation']
            }
    
    async def analyze_visual_elements_enhanced(self, video_frames: List[Dict]) -> Dict[str, Any]:
        """
        Enhanced visual analysis with more frames and detailed metrics
        """
        if not video_frames:
            return {'error': 'No video frames to analyze'}
        
        max_frames = ANALYSIS_CONFIG["sampling"]["max_frames_analyzed"]
        selected_frames = video_frames[:max_frames]
        
        logger.info(f"📊 Analyzing {len(selected_frames)} frames with enhanced visual metrics")
        await self.progress_callback(self.analysis_id, 60, f"📊 Analyzing {len(selected_frames)} frames with enhanced visual metrics")
        
        frame_analyses = []
        
        for i, frame_data in enumerate(selected_frames):
            frame = frame_data['frame']
            timestamp = frame_data['timestamp']

            # Calculate progress for this frame within 60-75% range
            frame_progress = 60 + int((i / len(selected_frames)) * 15)
            
            logger.info(f"📊 Analyzing frame {i+1}/{len(selected_frames)} (t={timestamp:.1f}s)")
            await self.progress_callback(
                self.analysis_id,
                frame_progress,
                f"📊 Analyzing frame {i+1}/{len(selected_frames)} (t={timestamp:.1f}s)"
            )
                   
            # Convert frame to base64 for OpenAI Vision API
            _, buffer = cv2.imencode('.jpg', frame)
            frame_base64 = base64.b64encode(buffer).decode('utf-8')
            
            # Enhanced frame analysis with GPT-4o Vision
            try:
                response = openai_client.chat.completions.create(
                    model="gpt-4o", # Vision requires gpt-4o, not mini
                    messages=[
                        {
                            "role": "user",
                            "content": [
                                {
                                    "type": "text",
                                    "text": f"""Analyze this lecture frame (timestamp: {timestamp:.1f}s) for detailed pedagogical elements:
                                    
                                    1. Eye Contact & Gaze Direction: 
                                       - Looking directly at camera/students vs reading/looking away
                                       - Natural vs forced eye contact
                                       - Consistency of gaze engagement
                                    
                                    2. Hand Gestures & Body Language:
                                       - Purposeful gestures supporting content vs nervous/distracting movements
                                       - Open, engaging gestures vs closed, defensive postures
                                       - Gesture variety and naturalness
                                    
                                    3. Posture & Positioning:
                                       - Confident, upright stance vs slouching/poor posture
                                       - Appropriate positioning relative to camera/audience
                                       - Movement and spatial awareness
                                    
                                    4. Facial Expressions & Engagement:
                                       - Animated, engaging expressions vs flat/monotone
                                       - Appropriate emotional expression for content
                                       - Genuine enthusiasm and interest
                                    
                                    5. Professional Appearance:
                                       - Appropriate dress and grooming
                                       - Visual presentation quality
                                       - Overall professional demeanor
                                    
                                    Rate each aspect from 1-10 and provide specific observations.
                                    Return as JSON with keys: eye_contact_score, gestures_score, posture_score, engagement_score, professionalism_score, detailed_observations"""
                                },
                                {
                                    "type": "image_url",
                                    "image_url": {
                                        "url": f"data:image/jpeg;base64,{frame_base64}",
                                        "detail": "high"
                                    }
                                }
                            ]
                        }
                    ],
                    max_completion_tokens=600
                )
                
                # Check if response and choices exist
                if not response or not response.choices or len(response.choices) == 0:
                    raise ValueError("OpenAI API returned empty response or no choices")
                
                # Check if response content exists
                response_content = response.choices[0].message.content
                if not response_content:
                    raise ValueError("AI response content is None or empty")
                
                analysis = self._safe_json_loads(response_content)
                analysis['timestamp'] = timestamp
                frame_analyses.append(analysis)
                logger.info(f"📊 Frame {i+1}/{len(selected_frames)} analyzed (t={timestamp:.1f}s)")
            except (json.JSONDecodeError, ValueError, AttributeError, APIError, APIConnectionError, APITimeoutError, RateLimitError, Exception) as e:
                # Enhanced fallback with timestamp
                error_msg = str(e) if e else "Unknown error"
                error_type = type(e).__name__
                logger.warning(f"⚠️ Frame {i+1} analysis failed ({error_type}): {error_msg}, using fallback scores")
                frame_analyses.append({
                    'eye_contact_score': 7,
                    'gestures_score': 7,
                    'posture_score': 7,
                    'engagement_score': 7,
                    'professionalism_score': 8,
                    'detailed_observations': [f'Unable to parse analysis for frame at {timestamp:.1f}s: {error_msg}'],
                    'timestamp': timestamp
                })
        
        # Enhanced aggregation with temporal weighting
        if frame_analyses:
            # Calculate weighted averages (give more weight to middle frames)
            total_frames = len(frame_analyses)
            
            weighted_scores = {
                'eye_contact': 0,
                'gestures': 0,
                'posture': 0,
                'engagement': 0,
                'professionalism': 0
            }
            
            total_weight = 0
            
            for i, analysis in enumerate(frame_analyses):
                # Weight frames in the middle more heavily (bell curve weighting)
                position_ratio = i / (total_frames - 1) if total_frames > 1 else 0.5
                weight = 1.0 - abs(position_ratio - 0.5) * 0.5  # Ranges from 0.75 to 1.0
                
                weighted_scores['eye_contact'] += analysis.get('eye_contact_score', 7) * weight
                weighted_scores['gestures'] += analysis.get('gestures_score', 7) * weight
                weighted_scores['posture'] += analysis.get('posture_score', 7) * weight
                weighted_scores['engagement'] += analysis.get('engagement_score', 7) * weight
                weighted_scores['professionalism'] += analysis.get('professionalism_score', 8) * weight
                
                total_weight += weight
            
            # Normalize by total weight
            for key in weighted_scores:
                weighted_scores[key] = weighted_scores[key] / total_weight if total_weight > 0 else 7.0
            
            # Collect all detailed observations
            all_observations = []
            for analysis in frame_analyses:
                obs = analysis.get('detailed_observations', [])
                if isinstance(obs, list):
                    all_observations.extend(obs)
                elif isinstance(obs, str):
                    all_observations.append(obs)
            
            return {
                'scores': weighted_scores,
                'observations': all_observations,
                'frames_analyzed': len(frame_analyses),
                'temporal_analysis': {
                    'frame_timestamps': [f['timestamp'] for f in frame_analyses],
                    'score_progression': {
                        'eye_contact': [f.get('eye_contact_score', 7) for f in frame_analyses],
                        'gestures': [f.get('gestures_score', 7) for f in frame_analyses],
                        'engagement': [f.get('engagement_score', 7) for f in frame_analyses]
                    }
                }
            }
        
        return {'error': 'No frames successfully analyzed'}
    
    def _ensure_mars_pedagogy_fields(self, p: Dict[str, Any]) -> Dict[str, Any]:
        """Ensure MARS v20260224 nine content criteria exist; map legacy five-dimension scores if needed."""
        co = float(p.get("content_organization") or 7.0)
        cc = float(p.get("communication_clarity") or 7.0)
        ue = float(p.get("use_of_examples") or 7.0)
        defaults = {
            "structural_sequencing": co,
            "logical_consistency": co,
            "closure_framing": co,
            "conceptual_accuracy": cc,
            "causal_reasoning_depth": cc,
            "multi_perspective_explanation": cc,
            "example_quality_frequency": ue,
            "analogy_concept_bridging": ue,
            "representation_diversity": ue,
        }
        for k, v in defaults.items():
            if k not in p or p[k] is None:
                p[k] = v
        # Evidence strings for UI "Why this score" (optional; filled by LLM)
        for k in defaults.keys():
            ek = f"evidence_{k}"
            if ek not in p or p[ek] is None:
                p[ek] = ""
        return p

    def _augment_causal_reasoning_depth(self, transcript: str, llm_score: float) -> float:
        """Blend LLM score with causal-marker density (Excel: causal connectors)."""
        if not transcript or not transcript.strip():
            return round(min(10.0, max(0.0, llm_score)), 1)
        t = transcript.lower()
        markers = (
            "therefore", "because", "thus", "hence", "as a result", "due to", "consequently",
            "cause ", " led to", "resulting in", "for this reason", "since ", "so that ",
        )
        count = sum(t.count(m) for m in markers)
        words = max(len(transcript.split()), 1)
        density = min(1.0, (count / words) * 500.0)
        heuristic = 3.0 + density * 7.0
        blended = 0.78 * float(llm_score) + 0.22 * heuristic
        return round(min(10.0, max(0.0, blended)), 1)

    def _extract_json_object_string(self, raw: str) -> str:
        """
        Extract the first top-level JSON object from LLM text (balanced braces, string-aware).
        Avoids naive first-{ to last-} slicing which breaks on nested structures or extra prose.
        """
        if raw is None:
            raise ValueError("empty")
        txt = str(raw).strip()
        if not txt:
            raise ValueError("empty")
        if txt.startswith("```"):
            txt = re.sub(r"^```[a-zA-Z0-9_-]*\s*", "", txt)
            txt = re.sub(r"\s*```\s*$", "", txt)
            txt = txt.strip()
        start = txt.find("{")
        if start < 0:
            raise ValueError("no json object start")
        depth = 0
        in_str = False
        esc = False
        for i in range(start, len(txt)):
            ch = txt[i]
            if esc:
                esc = False
                continue
            if in_str:
                if ch == "\\":
                    esc = True
                elif ch == '"':
                    in_str = False
                continue
            if ch == '"':
                in_str = True
                continue
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    return txt[start : i + 1]
        raise ValueError("unbalanced json object")

    def _safe_json_loads(self, raw: str) -> Dict[str, Any]:
        """
        Best-effort JSON extraction for LLM outputs.
        Handles code fences, surrounding prose, and mild trailing-comma issues.
        """
        blob = self._extract_json_object_string(raw)
        try:
            return json.loads(blob)
        except json.JSONDecodeError:
            blob2 = re.sub(r",\s*}", "}", blob)
            blob2 = re.sub(r",\s*]", "]", blob2)
            return json.loads(blob2)

    def _chat_json_completion(
        self,
        messages: List[Dict[str, str]],
        max_completion_tokens: int,
        prefer_json_object_format: bool = False,
    ):
        """Chat completion; optionally request JSON mode, with one fallback if the API rejects it."""
        kwargs: Dict[str, Any] = {
            "model": "gpt-5-nano",
            "messages": messages,
            "max_completion_tokens": max_completion_tokens,
        }
        if prefer_json_object_format:
            kwargs["response_format"] = {"type": "json_object"}
            try:
                return openai_client.chat.completions.create(**kwargs)
            except Exception:
                kwargs.pop("response_format", None)
                return openai_client.chat.completions.create(**kwargs)
        return openai_client.chat.completions.create(**kwargs)

    def _snippet_around(self, text: str, needle: str, radius: int = 110) -> str:
        if not text:
            return ""
        t = text
        idx = t.lower().find((needle or "").lower())
        if idx < 0:
            return ""
        lo = max(0, idx - radius)
        hi = min(len(t), idx + len(needle) + radius)
        snip = t[lo:hi].replace("\n", " ").strip()
        if lo > 0:
            snip = "…" + snip
        if hi < len(t):
            snip = snip + "…"
        return snip

    def _fill_mars_content_evidence_fallback(self, transcript: str, p: Dict[str, Any], lecture_context: str = "") -> Dict[str, Any]:
        """
        Populate evidence_<criterion> when the LLM doesn't provide useful evidence.
        Uses concrete transcript snippets and simple quantitative cues (keyword counts).
        """
        t = (transcript or "").strip()
        if not t:
            return p

        lc = (lecture_context or "").strip()
        tl = t.lower()
        context_line = ""
        if lc:
            # Use the already computed alignment signal if available; else fall back to cheap overlap check.
            alignment = p.get("context_alignment_score", None)
            verdict = (p.get("context_alignment_verdict") or "").strip()
            rationale = (p.get("context_alignment_rationale") or "").strip()
            try:
                alignment_f = float(alignment) if alignment is not None else None
            except Exception:
                alignment_f = None

            # Pick up to 3 meaningful context keywords and try to find a matching snippet in transcript.
            stop = {
                "this","that","these","those","the","a","an","and","or","but","to","of","in","on","for","with","as","at","by","from",
                "is","are","was","were","be","been","being","it","we","you","they","i","our","your","their",
                "lecture","session","week","module","course","topic","learning","outcome","outcomes","students","student","audience",
            }
            toks = re.findall(r"[a-zA-Z][a-zA-Z0-9_-]{2,}", lc.lower())[:220]
            uniq = []
            for x in toks:
                if x in stop:
                    continue
                if x not in uniq:
                    uniq.append(x)
                if len(uniq) >= 12:
                    break
            hits = [w for w in uniq if w in tl]
            snippet = ""
            if hits:
                for w in hits[:3]:
                    snippet = self._snippet_around(t, w)
                    if snippet:
                        break
            # If nothing matches, flag low overlap.
            if alignment_f is None:
                overlap = len(hits) / max(1, len(uniq))
                alignment_f = overlap
                verdict = verdict or ("match" if overlap >= 0.35 else ("partial" if overlap >= 0.15 else "mismatch"))
            verdict_norm = verdict.lower() if verdict else ""
            if verdict_norm in ("mismatch", "off", "off-focus") or (alignment_f is not None and alignment_f <= 0.25):
                context_line = (
                    "Context alignment appears weak: the submitted lecture context does not strongly match the transcript excerpt. "
                    + (f"Example transcript cue: \"{snippet}\". " if snippet else "")
                    + (f"Model rationale: {rationale} " if rationale else "")
                    + "This may reduce Content scores because the structure/explanations are evaluated against the stated module/topic."
                ).strip()
            else:
                context_line = (
                    "Context alignment appears reasonable: the transcript contains cues consistent with the submitted lecture context. "
                    + (f"Example cue: \"{snippet}\". " if snippet else "")
                    + (f"Model rationale: {rationale} " if rationale else "")
                ).strip()

        words = max(1, len(t.split()))

        def per_1k(n: int) -> float:
            return round((n / words) * 1000.0, 1)

        seq_markers = ("first", "second", "third", "next", "then", "finally", "to begin", "in summary")
        seq_count = sum(t.lower().count(m) for m in seq_markers)

        recap_markers = ("in summary", "to summarise", "to conclude", "key takeaway", "the main point", "recap")
        recap_count = sum(t.lower().count(m) for m in recap_markers)

        ex_markers = ("for example", "for instance", "e.g.", "let's take", "consider")
        ex_count = sum(t.lower().count(m) for m in ex_markers)

        analogy_markers = ("like ", "similar to", "think of", "as if", "imagine")
        analogy_count = sum(t.lower().count(m) for m in analogy_markers)

        contrast_markers = ("on the other hand", "alternatively", "compare", "in contrast", "whereas")
        contrast_count = sum(t.lower().count(m) for m in contrast_markers)

        causal_markers = ("because", "therefore", "thus", "as a result", "due to", "hence", "consequently")
        causal_count = sum(t.lower().count(m) for m in causal_markers)

        def set_if_empty(key: str, val: str):
            ek = f"evidence_{key}"
            if not str(p.get(ek, "") or "").strip():
                p[ek] = (val or "").strip()

        set_if_empty(
            "structural_sequencing",
            f"Evidence: sequencing markers appear ~{per_1k(seq_count)} per 1,000 words (e.g., first/next/then/finally). "
            f"Example: \"{self._snippet_around(t, 'first') or self._snippet_around(t, 'next') or self._snippet_around(t, 'then')}\"",
        )
        set_if_empty(
            "logical_consistency",
            "Evidence: within the excerpt provided, explanations reuse terms and link ideas without obvious contradictions. "
            "If this is a highly technical lecture, verify key claims against your module materials; the system cannot fully validate every domain-specific statement.",
        )
        set_if_empty(
            "closure_framing",
            f"Evidence: recap/closure cues appear ~{per_1k(recap_count)} per 1,000 words (e.g., in summary/to conclude/recap). "
            f"Example: \"{self._snippet_around(t, 'in summary') or self._snippet_around(t, 'to conclude') or self._snippet_around(t, 'recap')}\"",
        )
        set_if_empty(
            "conceptual_accuracy",
            "Evidence: the score is inferred from how consistently concepts are defined and used in the transcript excerpt. "
            "For strong assurance, cross-check key definitions in the transcript against your slides/notes (the system may miss subtle technical inaccuracies).",
        )
        set_if_empty(
            "causal_reasoning_depth",
            f"Evidence: causal connectors appear ~{per_1k(causal_count)} per 1,000 words (e.g., because/therefore/as a result), indicating 'why/how' reasoning. "
            f"Example: \"{self._snippet_around(t, 'because') or self._snippet_around(t, 'therefore') or self._snippet_around(t, 'as a result')}\"",
        )
        set_if_empty(
            "multi_perspective_explanation",
            f"Evidence: comparison/contrast cues appear ~{per_1k(contrast_count)} per 1,000 words (e.g., on the other hand/alternatively/compare). "
            f"Example: \"{self._snippet_around(t, 'on the other hand') or self._snippet_around(t, 'alternatively') or self._snippet_around(t, 'compare')}\"",
        )
        set_if_empty(
            "example_quality_frequency",
            f"Evidence: example cues appear ~{per_1k(ex_count)} per 1,000 words (e.g., for example/for instance/consider). "
            f"Example: \"{self._snippet_around(t, 'for example') or self._snippet_around(t, 'for instance') or self._snippet_around(t, 'consider')}\"",
        )
        set_if_empty(
            "analogy_concept_bridging",
            f"Evidence: analogy cues appear ~{per_1k(analogy_count)} per 1,000 words (e.g., think of/similar to/imagine). "
            f"Example: \"{self._snippet_around(t, 'think of') or self._snippet_around(t, 'similar to') or self._snippet_around(t, 'imagine')}\"",
        )
        set_if_empty(
            "representation_diversity",
            "Evidence: representation diversity is inferred from references to multiple forms (e.g., verbal explanation plus equations/diagrams/slides). "
            "If the recording is slides-heavy without verbal description, the system may under-detect representations that are only visible on screen.",
        )

        # Attach context line to all nine criteria evidence blocks (Content 1.1–1.3) if context exists.
        if context_line:
            for k in (
                "structural_sequencing",
                "logical_consistency",
                "closure_framing",
                "conceptual_accuracy",
                "causal_reasoning_depth",
                "multi_perspective_explanation",
                "example_quality_frequency",
                "analogy_concept_bridging",
                "representation_diversity",
            ):
                ek = f"evidence_{k}"
                cur = str(p.get(ek, "") or "").strip()
                if cur and context_line.lower() not in cur.lower():
                    p[ek] = f"{cur} {context_line}".strip()
        return p

    async def analyze_student_feedback_metrics(self, speech_analysis: Dict) -> Dict[str, Any]:
        """
        MARS Engagement → Feedback: learner question frequency & cognitive level (0–10 each).
        Webcasts often lack audience audio — LLM estimates confidence; confidence `none` → neutral 5/10 each + remarks (avoids double-penalising with instructor-only recordings).
        """
        transcript = speech_analysis.get("transcript") or ""
        if len(transcript.strip()) < 80:
            return {
                "student_question_frequency_score": 0.0,
                "student_question_cognitive_score": 0.0,
                "student_feedback_confidence": "none",
                "student_feedback_remarks": "Transcript too short to assess audience questions.",
                "audience_questions": [],
                "audience_question_count": 0,
            }
        try:
            response = openai_client.chat.completions.create(
                model="gpt-5-nano",
                messages=[
                    {
                        "role": "system",
                        "content": """You assess whether the transcript contains student/audience questions (not the instructor's questions).
In typical webcasts only the instructor is mic'd: if all questions appear to be instructor questions, set scores to 0 and confidence none.
Return JSON only:
- student_question_frequency_score: number 0-10 (density of clear audience/student questions)
- student_question_cognitive_score: number 0-10 (Bloom-style depth of those questions)
- confidence: one of none|low|medium|high
- remarks: short note (e.g. no separate audience audio)
- audience_questions: array of objects {{"question": string, "context": string}} listing each distinct student/audience question you can identify (empty if none)
- audience_question_count: integer (length of audience_questions)""",
                    },
                    {
                        "role": "user",
                        "content": f"Transcript:\n{transcript[:12000]}",
                    },
                ],
                max_completion_tokens=800,
            )
            txt = response.choices[0].message.content
            if not txt:
                raise ValueError("empty")
            data = self._safe_json_loads(txt)
            conf = (data.get("confidence") or "none").lower()
            if conf == "none":
                # Webcast: no evidenced audience channel — use neutral 5/10 so MARS Engagement is not double-penalised
                # (instructor prompting density already carries interaction signal).
                data["student_question_frequency_score"] = 5.0
                data["student_question_cognitive_score"] = 5.0
            sf = float(data.get("student_question_frequency_score", 0) or 0)
            sc = float(data.get("student_question_cognitive_score", 0) or 0)
            aq = data.get("audience_questions") or []
            if not isinstance(aq, list):
                aq = []
            return {
                "student_question_frequency_score": round(min(10.0, max(0.0, sf)), 1),
                "student_question_cognitive_score": round(min(10.0, max(0.0, sc)), 1),
                "student_feedback_confidence": conf,
                "student_feedback_remarks": data.get("remarks", "") or (
                    "Student voice not evidenced in transcript; learner-question subscores set to neutral 5/10 (does not penalise engagement)."
                    if conf == "none" else ""
                ),
                "audience_questions": aq[:30],
                "audience_question_count": int(data.get("audience_question_count") if data.get("audience_question_count") is not None else len(aq)),
            }
        except Exception as e:
            logger.warning(f"Student feedback metrics fallback: {e}")
            return {
                "student_question_frequency_score": 5.0,
                "student_question_cognitive_score": 5.0,
                "student_feedback_confidence": "none",
                "student_feedback_remarks": "Student voice not evidenced; learner-question subscores default to neutral 5/10.",
                "audience_questions": [],
                "audience_question_count": 0,
            }

    def _normalise_question_key(self, q: str) -> str:
        """Normalise question text for matching (lowercase, collapse whitespace, strip punctuation)."""
        s = (q or "").strip().lower()
        s = re.sub(r"\s+", " ", s)
        s = re.sub(r"[^\w\s]", "", s)
        return s

    def _compute_student_uptake_index(self, transcript: str, sf_metrics: Dict[str, Any]) -> Dict[str, Any]:
        """
        SUI definition (revised): extent to which instructor builds upon/incorporates student responses.
        For webcasts without learner voice, return a conservative score and clear disclaimer.
        """
        conf = (sf_metrics.get("student_feedback_confidence") or "none").lower()
        t = (transcript or "").strip()
        if conf == "none":
            return {
                "score": 2.0,
                "evidence": "Student voice may not be recorded clearly in webcast lectures; uptake cannot be reliably determined from transcript alone. Scored conservatively (2/10).",
            }
        # Heuristic uptake cues (instructor referencing or building on learner input)
        cues = (
            "good question", "great question", "as you said", "as someone said", "as one of you said",
            "as you mentioned", "as mentioned earlier", "building on", "following up on",
            "to respond to", "answer to your question", "someone asked", "your question",
        )
        tl = t.lower()
        count = sum(tl.count(c) for c in cues)
        words = max(1, len(t.split()))
        per_1k = round((count / words) * 1000.0, 2)
        # Map cue density to 0–10 with a gentle slope; clamp and round.
        # 0 cues → 2.5; 1/1k → ~4; 3/1k → ~7; >=5/1k → 10
        score = 2.5 + min(7.5, per_1k * 2.5)
        score = round(min(10.0, max(0.0, score)), 1)
        ev = f"Evidence: uptake cues occur ~{per_1k:.2f} per 1,000 words (e.g., “good question”, “building on …”, “to respond to …”)."
        return {"score": score, "evidence": ev}

    async def analyze_pedagogy_enhanced(
        self, speech_analysis: Dict, visual_analysis: Dict, lecture_context: str = ""
    ) -> Dict[str, Any]:
        """
        MARS Content category: nine criteria (Revised Rubric) + legacy five-dimension scores for reports.
        """
        lc = (lecture_context or "").strip()
        lc_block = (
            f"\n---\nINSTRUCTOR-PROVIDED LECTURE CONTEXT (authoritative for intended subject, course, level, and learning goals):\n{lc}\n---\n"
            if lc
            else "\n(No instructor lecture context was provided; you cannot judge topic alignment against stated course goals—score internal coherence only, and note the limitation in evidence where relevant.)\n"
        )

        context = f"""
        COMPREHENSIVE LECTURE ANALYSIS DATA:
        {lc_block}
        Speech Analysis (Full Transcript excerpt):
        - Transcript: {speech_analysis.get('transcript', '')[:6000]}...
        - Speaking Rate: {speech_analysis.get('speaking_rate', 0):.1f} WPM
        - Filler Word Ratio: {speech_analysis.get('filler_ratio', 0):.3f}
        - Key Highlights: {'; '.join(speech_analysis.get('highlights', [])[:8])}
        
        Visual summary ({visual_analysis.get('frames_analyzed', 0)} frames):
        - Scores — eye_contact: {visual_analysis.get('scores', {}).get('eye_contact', 'N/A')}, gestures: {visual_analysis.get('scores', {}).get('gestures', 'N/A')}, posture: {visual_analysis.get('scores', {}).get('posture', 'N/A')}, engagement: {visual_analysis.get('scores', {}).get('engagement', 'N/A')}, professionalism: {visual_analysis.get('scores', {}).get('professionalism', 'N/A')}
        """

        system_pedagogy = """You are an expert pedagogical analyst. Score the lecture transcript (1-10 each) for MARS Content rubric:

A) Content Organisation (MARS 1.1): structural_sequencing, logical_consistency, closure_framing
B) Explanation Quality (MARS 1.2): conceptual_accuracy, causal_reasoning_depth, multi_perspective_explanation
C) Use of Examples / Representation (MARS 1.3): example_quality_frequency, analogy_concept_bridging, representation_diversity

CRITICAL — LECTURE CONTEXT AND TOPIC ALIGNMENT:
- When the instructor has provided "INSTRUCTOR-PROVIDED LECTURE CONTEXT", treat it as the ground truth for what this session should be about (e.g. course name, subject, topic, level, ILOs).
- For ALL nine criteria in A, B, and C above, you MUST evaluate whether the spoken content is appropriate and aligned with that context.
- If the transcript is largely about a different discipline or topic than the context states (e.g. extensive coding or unrelated content in a Japanese language class), the lecture is NOT instructionally successful for that course: score the nine criteria LOW (typically 2–5) and explain the misalignment explicitly in each evidence_* string. Do not give high scores for "clear structure" of the wrong subject matter.
- If context is missing, do not invent alignment; score based on internal coherence of the transcript and state in evidence that alignment to course goals could not be verified.

Also provide legacy aggregate scores (1-10): content_organization, engagement_techniques, communication_clarity, use_of_examples, knowledge_checking, overall_effectiveness.

Return a COMPACT JSON only response. Do NOT include long essays or multi-paragraph fields.
- strengths: 3 short bullet-like strings
- improvements: 3 short bullet-like strings
- recommendations: 3 short bullet-like strings
- alignment_comment: 1-2 sentences noting whether transcript matches the instructor-provided context (or that context is missing).

You must output JSON only (one object) so it can be parsed programmatically."""

        user_pedagogy = f"""{context}

Return a single JSON object with ALL keys:
structural_sequencing, logical_consistency, closure_framing,
conceptual_accuracy, causal_reasoning_depth, multi_perspective_explanation,
example_quality_frequency, analogy_concept_bridging, representation_diversity,
content_organization, engagement_techniques, communication_clarity, use_of_examples, knowledge_checking, overall_effectiveness,
strengths, improvements, recommendations, alignment_comment"""

        p: Optional[Dict[str, Any]] = None
        last_err: Optional[Exception] = None
        for attempt in range(2):
            suffix = ""
            if attempt == 1:
                suffix = (
                    "\n\nIMPORTANT: Your previous answer could not be parsed. Reply with ONE valid JSON object only. "
                    "No markdown code fences and no text before or after the object. Use double-quoted keys and strings."
                )
            try:
                response = self._chat_json_completion(
                    messages=[
                        {"role": "system", "content": system_pedagogy},
                        {"role": "user", "content": user_pedagogy + suffix},
                    ],
                    max_completion_tokens=1800,
                    prefer_json_object_format=(attempt == 0),
                )
                if not response or not getattr(response, "choices", None):
                    raise ValueError("empty completion")
                msg = response.choices[0].message
                response_content = (getattr(msg, "content", None) or "").strip()
                if not response_content:
                    raise ValueError("AI response content is None or empty")
                p = self._safe_json_loads(response_content)
                break
            except Exception as e:
                last_err = e
                logger.warning("Pedagogy JSON attempt %s failed: %s", attempt + 1, e)

        if p is None:
            logger.error("Pedagogy analysis using template fallback after JSON failures: %s", last_err)
            fb = {
                # Neutral placeholders; do not surface parse errors as "strengths" in the user report.
                "content_organization": 5.0,
                "engagement_techniques": 5.0,
                "communication_clarity": 5.0,
                "use_of_examples": 5.0,
                "knowledge_checking": 5.0,
                "overall_effectiveness": 5.0,
                "strengths": [],
                "improvements": [],
                "recommendations": [],
                "alignment_comment": "Pedagogical content scoring is temporarily unavailable (response could not be parsed). Please re-run analysis.",
                "pedagogy_parse_failed": True,
            }
            fb = self._ensure_mars_pedagogy_fields(fb)
            fb["lecture_context_provided"] = bool(lc)
            return fb

        p = self._ensure_mars_pedagogy_fields(p)
        p["lecture_context_provided"] = bool(lc)
        p["pedagogy_parse_failed"] = False
        if lc:
            try:
                align_resp = self._chat_json_completion(
                    messages=[
                        {
                            "role": "system",
                            "content": (
                                "You judge topical alignment between instructor-provided lecture context and the transcript.\n"
                                "Return ONLY JSON with keys: alignment_score (number 0.0-1.0), verdict (one of: match, partial, mismatch), rationale (1-3 sentences).\n"
                                "alignment_score meaning: 1.0 = clearly the same topic/discipline; 0.0 = clearly different discipline/topic.\n"
                                "Be strict: if context says one discipline (e.g. political science) but transcript is mostly another (e.g. BIM / construction), verdict=mismatch and alignment_score <= 0.2."
                            ),
                        },
                        {
                            "role": "user",
                            "content": (
                                f"LECTURE CONTEXT:\n{lc}\n\n"
                                f"TRANSCRIPT EXCERPT:\n{(speech_analysis.get('transcript') or '')[:7000]}\n"
                            ),
                        },
                    ],
                    max_completion_tokens=400,
                    prefer_json_object_format=True,
                )
                ac = ""
                if align_resp and align_resp.choices:
                    ac = (align_resp.choices[0].message.content or "").strip()
                aj = self._safe_json_loads(ac) if ac else {}
                try:
                    p["context_alignment_score"] = float(aj.get("alignment_score"))
                except Exception:
                    p["context_alignment_score"] = None
                p["context_alignment_verdict"] = aj.get("verdict")
                p["context_alignment_rationale"] = aj.get("rationale")
            except Exception:
                p["context_alignment_score"] = None
                p["context_alignment_verdict"] = None
                p["context_alignment_rationale"] = None
        p["causal_reasoning_depth"] = self._augment_causal_reasoning_depth(
            speech_analysis.get("transcript") or "",
            float(p.get("causal_reasoning_depth") or 7.0),
        )
        p = self._fill_mars_content_evidence_fallback(speech_analysis.get("transcript") or "", p, lecture_context=lc)
        return p
        

    def _compute_question_distribution_stability(self, questions: List[Dict], duration_seconds: float) -> Dict[str, Any]:
        """
        Question Distribution Stability (QDS): split lecture duration into 5 equal time quintiles
        (0–20%, 20–40%, …). Each quintile that contains at least one question scores 2 points (max 10).
        """
        nq = len(questions) if questions else 0
        if nq <= 0 or duration_seconds <= 0:
            return {
                "score": 0.0,
                "n": nq,
                "cv": None,
                "mean_gap_s": None,
                "quintile_hits": [False] * 5,
                "quintiles_filled": 0,
            }
        D = float(duration_seconds)
        quintile_hits = [False] * 5
        for q in questions:
            t = float(q.get("start_time") or 0)
            if t < 0:
                t = 0.0
            frac = t / D
            if frac >= 1.0:
                idx = 4
            else:
                idx = int(frac * 5)
                idx = min(4, max(0, idx))
            quintile_hits[idx] = True
        k = sum(1 for h in quintile_hits if h)
        score = round(min(10.0, max(0.0, 2.0 * k)), 1)
        return {
            "score": score,
            "n": nq,
            "cv": None,
            "mean_gap_s": None,
            "quintile_hits": quintile_hits,
            "quintiles_filled": k,
        }

    def _compute_student_uptake_index_from_questions(
        self,
        words_data: List[Dict[str, Any]],
        questions: List[Dict[str, Any]],
        sf_metrics: Dict[str, Any],
    ) -> Dict[str, Any]:
        """
        Student Uptake Index (SUI): does the instructor build on learner responses?

        Revised logic:
        - If webcast has no reliable learner voice (confidence none), return conservative 2/10 + limitation note.
        - Otherwise, for each instructor question, inspect a short window AFTER the question (word indices) for
          uptake cues such as \"to answer your question\" / \"good question\" / \"as you said\".
        - Score = 2 + 8*(uptake_hits / total_questions), clamped 0-10.
        """
        conf = (sf_metrics.get("student_feedback_confidence") or "none").lower()
        if conf == "none":
            return {
                "score": 2.0,
                "uptake_hits": 0,
                "uptake_rate": 0.0,
                "evidence": (
                    "Student voice may not be recorded or distinguishable in webcast lectures; "
                    "uptake cannot be reliably determined from transcript alone. Scored conservatively (2/10)."
                ),
            }
        cues = (
            "good question",
            "great question",
            "to answer your question",
            "answer your question",
            "someone asked",
            "as you said",
            "as someone said",
            "building on",
            "following up on",
            "to respond to",
        )
        total = int(len(questions) or 0)
        if total <= 0:
            return {"score": 0.0, "uptake_hits": 0, "uptake_rate": 0.0, "evidence": "No instructor questions to evaluate uptake around."}
        hits = 0
        samples = []
        for q in questions[:120]:
            end_idx = q.get("end_idx")
            if end_idx is None:
                continue
            try:
                end_i = int(end_idx)
            except Exception:
                continue
            window = words_data[end_i : min(len(words_data), end_i + 65)]
            tail = " ".join((w.get("word") or "") for w in window).strip().lower()
            if not tail:
                continue
            if any(c in tail for c in cues):
                hits += 1
                if len(samples) < 2:
                    samples.append(" / ".join([c for c in cues if c in tail][:2]))
        rate = hits / float(total) if total else 0.0
        score = round(min(10.0, max(0.0, 2.0 + 8.0 * rate)), 1)
        ev = f"Detected uptake cues after {hits}/{total} instructor questions (rate {rate:.2f})."
        if samples:
            ev += f" Example cues: {', '.join(samples)}."
        return {"score": score, "uptake_hits": hits, "uptake_rate": round(rate, 3), "evidence": ev}

    def _sui_from_prompting_density(
        self,
        duration_minutes: float,
        total_questions: int,
        count_active: int,
        count_constructive: int,
        count_interactive: int,
    ) -> float:
        """
        Webcast-fair proxy for learner engagement opportunity: frequent Active / C / I prompts
        create response slots even when audience audio is missing from the recording.
        """
        dm = max(0.1, float(duration_minutes))
        qpm = total_questions / dm
        eqd = (count_constructive + count_interactive) / dm
        apd = count_active / dm
        s = 3.0 + 2.35 * eqd + 0.88 * apd
        if qpm >= 2.5:
            s += 1.35
        elif qpm >= 2.0:
            s += 1.05
        elif qpm >= 1.5:
            s += 0.65
        elif qpm >= 1.0:
            s += 0.35
        return round(min(10.0, max(0.0, s)), 1)

    def _compute_question_quality_from_icap_counts(
        self,
        total_questions: int,
        count_passive: int,
        count_active: int,
        count_constructive: int,
        count_interactive: int,
        _duration_minutes: float,
    ) -> Tuple[float, float, str]:
        """
        CLI (3.2.1): primary driver is % of questions classified as Interactive.
        >20% → 9/10; 10%–20% (inclusive) → 7/10; 5%–<10% → 5/10; below 5% → 3/10.
        Small bump when Constructive share is high (does not override Interactive band).
        """
        if total_questions <= 0:
            return 0.0, 0.0, "low"
        tq = float(total_questions)
        pi = count_interactive / tq
        pc = count_constructive / tq

        if pi > 0.20:
            base = 9.0
        elif pi >= 0.10:
            base = 7.0
        elif pi >= 0.05:
            base = 5.0
        else:
            base = 3.0

        if pc >= 0.25:
            base = min(10.0, base + 0.5)
        elif pc >= 0.15:
            base = min(10.0, base + 0.25)

        if count_passive == total_questions:
            base = min(base, 4.0)

        question_quality = round(min(10.0, max(1.0, base)), 1)
        cli_100 = round(question_quality * 10.0, 1)
        if question_quality >= 8.5:
            cog = "high"
        elif question_quality >= 6.0:
            cog = "medium"
        else:
            cog = "low"
        return question_quality, cli_100, cog

    def _question_engagement_narrative(
        self,
        total_questions: int,
        duration_minutes: float,
        count_active: int,
        count_interactive: int,
        count_constructive: int,
        high_level_count: int,
    ) -> str:
        dm = max(0.1, float(duration_minutes))
        qpm = round(total_questions / dm, 2)
        apm = round(count_active / dm, 2)
        ipm = round(count_interactive / dm, 2)
        cpm = round(count_constructive / dm, 2)
        return (
            f"Across about {dm:.1f} minutes, the scan found {total_questions} instructor questions "
            f"({qpm}/min). By ICAP, {count_active} are Active (≈{apm}/min), {count_interactive} Interactive (≈{ipm}/min), "
            f"{count_constructive} Constructive (≈{cpm}/min). "
            f"{high_level_count} questions are Constructive or Interactive (higher-order prompts); "
            "those contribute extra weight in the engagement-quality score."
        )

    def _normalise_icap_label(self, raw: Optional[str]) -> Optional[str]:
        if not raw or not str(raw).strip():
            return None
        x = str(raw).strip().lower()
        if x == "passive":
            return "Passive"
        if x == "active":
            return "Active"
        if x == "constructive":
            return "Constructive"
        if x == "interactive":
            return "Interactive"
        return None

    def _classify_icap_heuristic(self, text: str) -> str:
        """
        Rule-based ICAP when the LLM over-uses Active or mismatches transcript wording.
        Prefer Passive for tag questions and ultra-short non-interrogative ?-lines (ASR noise).
        """
        t = (text or "").strip()
        if not t:
            return "Passive"
        low = t.lower()
        words = re.findall(r"\b[\w']+\b", low)
        interrogatives = (
            "what", "when", "where", "who", "which", "why", "how", "whose", "whom",
        )
        has_int = any(w in interrogatives for w in words)

        # Passive: rhetorical / check-ins / tag questions
        passive_res = [
            r"^(right|okay|ok|yes|no|see|clear|alright|hmm|uh)\s*\??$",
            r"^(make sense|got it|you see|understand|everyone good|everyone okay)\b",
            r"\b(isn'?t it|aren'?t they|isn'?t that|right|okay|ok)\s*\?$",
            r"^any\s+questions?\b",
            r"\bany questions?\s*(so far|left|remaining|before we)\b",
            r"^(everyone|anybody|anyone)\s+(clear|good|with me)\b",
        ]
        for p in passive_res:
            if re.search(p, low):
                return "Passive"

        if len(words) <= 4 and not has_int:
            return "Passive"
        if len(words) <= 8 and not has_int:
            return "Passive"

        # Interactive: dialogue / peer / audience-oriented
        if re.search(
            r"\b(you|your|y'all|we all|would you|could you|do you agree|discuss with|turn to|with a partner)\b",
            low,
        ):
            if has_int or re.search(r"\b(why|how|what if|compare)\b", low):
                return "Interactive"
        if re.search(r"\b(what would you|how would you|what do you think|your view|opinion)\b", low):
            return "Interactive"

        # Constructive: reasoning / explanation prompts
        if re.search(
            r"\b(why|how come|explain|justify|what causes|what is the reason|evaluate|analyse|analyze|"
            r"compare and contrast|what happens if|infer|implications)\b",
            low,
        ):
            return "Constructive"

        # Active: factual recall
        if has_int or re.search(
            r"\b(is|are|was|were|does|do|did|can|could|should|will|has|have)\b", low
        ):
            return "Active"

        return "Passive"

    def _merge_icap_labels(self, llm: Optional[str], heuristic: str) -> str:
        """Combine LLM label with heuristic; reduce spurious Active and recover deeper types."""
        h = heuristic or "Active"
        if llm is None:
            return h
        # LLM often marks mumbling/tag lines as Active
        if h == "Passive" and llm == "Active":
            return "Passive"
        # Real deeper questions the heuristic missed
        if h == "Passive" and llm in ("Constructive", "Interactive"):
            return llm
        if llm == "Active" and h in ("Constructive", "Interactive"):
            return h
        if h == "Active" and llm in ("Constructive", "Interactive", "Passive"):
            return llm
        rank = {"Passive": 0, "Active": 1, "Constructive": 2, "Interactive": 3}
        return llm if rank.get(llm, 1) >= rank.get(h, 1) else h

    def _find_llm_icap_for_question(
        self,
        analyzed_list: List[Dict[str, Any]],
        q_text: str,
        index: int,
    ) -> Optional[str]:
        """Match LLM output to a detected question (exact key, fuzzy, or index fallback)."""
        if not analyzed_list:
            return None
        key = self._normalise_question_key(q_text)
        best_partial: Optional[str] = None
        for aq in analyzed_list:
            aq_text = (aq.get("question") or "").strip()
            ak = self._normalise_question_key(aq_text)
            if key and ak and key == ak:
                return self._normalise_icap_label(aq.get("icap"))
            if key and ak and len(key) >= 18 and (key in ak or ak in key):
                best_partial = self._normalise_icap_label(aq.get("icap"))
        if best_partial:
            return best_partial
        if index < len(analyzed_list):
            return self._normalise_icap_label(analyzed_list[index].get("icap"))
        return None

    def _should_exclude_question_sentence(self, sentence: str) -> bool:
        """
        Drop likely false positives: very short non-interrogative ?-fragments and pure tag lines.
        """
        s = (sentence or "").strip()
        if not s.endswith("?"):
            return True
        low = s.lower()
        words = re.findall(r"\b[\w']+\b", low)
        interrogatives = (
            "what", "when", "where", "who", "which", "why", "how", "whose", "whom",
        )
        has_int = any(w in interrogatives for w in words)

        if len(words) <= 2:
            return True
        if len(words) <= 4 and not has_int:
            return True
        # Standalone tag / filler questions
        if re.match(
            r"^(right|okay|ok|yes|no|see|clear|hmm|uh|really|sure)\s*\?$",
            low,
        ):
            return True
        return False

    def detect_questions_pattern_matching(self, words_data: List[Dict], transcript_text: str = "") -> List[Dict]:
        """
        Detect ALL questions from polished transcript.
        Uses polished transcript (with proper punctuation from GPT post-processing).
        
        Criteria for a valid question:
        1. Must end with a question mark (?)
        2. Must be a complete sentence (starts with capital letter, minimum 3 words)
        3. Includes ALL questions regardless of cognitive level or question word
        
        This function lists ALL questions first, then AI will analyze which are high-level.
        """
        if not transcript_text:
            logger.warning("⚠️ No polished transcript text provided for question detection")
            return []
        
        # Ensure all word_data items are dictionaries (convert TranscriptionWord objects if needed)
        converted_words_data = []
        for word_item in words_data:
            if hasattr(word_item, 'word'):
                # It's a TranscriptionWord object - convert to dict
                converted_words_data.append({
                    'word': word_item.word,
                    'start': getattr(word_item, 'start', 0),
                    'end': getattr(word_item, 'end', 0)
                })
            elif isinstance(word_item, dict):
                converted_words_data.append(word_item)
            else:
                # Fallback
                converted_words_data.append({
                    'word': word_item.get('word', '') if isinstance(word_item, dict) else getattr(word_item, 'word', ''),
                    'start': word_item.get('start', 0) if isinstance(word_item, dict) else getattr(word_item, 'start', 0),
                    'end': word_item.get('end', 0) if isinstance(word_item, dict) else getattr(word_item, 'end', 0)
                })
        
        words_data = converted_words_data
        
        detected_questions = []
        
        # Split transcript into sentences - use a simpler, more reliable approach
        import re
        
        # First, find all sentences ending with question marks using regex
        # Pattern: Start with capital letter, capture everything until question mark
        question_pattern = r'([A-Z][^.!?]*\?)'
        question_matches = re.findall(question_pattern, transcript_text)
        
        sentence_list = []
        for match in question_matches:
            sentence = match.strip()
            if sentence:
                sentence_list.append(sentence)
        
        # Also try splitting by sentence boundaries for completeness
        # Split on sentence-ending punctuation followed by space and capital letter
        sentences_split = re.split(r'([.!?])\s+([A-Z])', transcript_text)
        
        # Reconstruct sentences from split
        current_sentence = ""
        for i, part in enumerate(sentences_split):
            if i == 0:
                current_sentence = part
            elif i % 3 == 1:  # Punctuation mark
                current_sentence += part
            elif i % 3 == 2:  # Capital letter (start of next sentence)
                current_sentence += " " + part
                if current_sentence.strip().endswith('?'):
                    sentence = current_sentence.strip()
                    if sentence and sentence not in sentence_list:
                        sentence_list.append(sentence)
                current_sentence = ""
            else:
                current_sentence += part
        
        # Add last sentence if it ends with ?
        if current_sentence.strip().endswith('?'):
            sentence = current_sentence.strip()
            if sentence and sentence not in sentence_list:
                sentence_list.append(sentence)
        
        # Remove duplicates while preserving order
        seen = set()
        unique_sentences = []
        for s in sentence_list:
            s_lower = s.lower().strip()
            if s_lower not in seen:
                seen.add(s_lower)
                unique_sentences.append(s)
        sentence_list = unique_sentences
        
        # Filter sentences to find ALL questions (ending with ?)
        for sentence in sentence_list:
            sentence = sentence.strip()
            
            # Skip empty sentences
            if not sentence:
                continue
            
            # Criterion 1: Must end with question mark
            if not sentence.endswith('?'):
                continue
            
            # Criterion 2: Must start with a capital letter (complete sentence)
            if not sentence[0].isupper():
                continue
            
            # Criterion 3: Must have minimum length (at least 3 words)
            words = sentence.split()
            if len(words) < 3:
                continue
            
            # Criterion 4: Additional filtering - exclude very short fragments
            # Remove punctuation for length check
            sentence_clean = re.sub(r'[^\w\s]', '', sentence)
            if len(sentence_clean.split()) < 3:
                continue
            
            if self._should_exclude_question_sentence(sentence):
                continue
            
            # Find the corresponding word timestamps for this sentence
            # Match by finding the first few words in the word data
            start_time = None
            start_idx = None
            end_idx = None
            
            if words_data:
                # Try to find the sentence in word data by matching first 2-3 words
                first_words_lower = [w.lower().strip('.,!?;:') for w in words[:3]]
                
                for i in range(len(words_data) - len(first_words_lower) + 1):
                    # Get a window of words to match against
                    window_words = []
                    for j in range(i, min(i + 5, len(words_data))):
                        word = words_data[j].get('word', '').lower().strip('.,!?;:')
                        window_words.append(word)
                    
                    window_text = ' '.join(window_words)
                    
                    # Check if first 2-3 words of the sentence match the window
                    match_count = 0
                    for fw in first_words_lower:
                        if fw in window_text:
                            match_count += 1
                    
                    # If at least 2 words match, consider it a match
                    if match_count >= 2:
                        start_time = words_data[i].get('start', 0)
                        start_idx = i
                        # Estimate end index based on sentence length
                        end_idx = min(i + len(words), len(words_data))
                        break
            
            # If we found a match (or if no word data available), add the question
            if start_time is not None or not words_data:
                detected_questions.append({
                    'question': sentence,
                    'start_time': start_time if start_time is not None else 0,
                    'start_idx': start_idx,
                    'end_idx': end_idx,
                    'confidence': 'high',
                    'detection_method': 'all_questions_with_question_mark'
                })
        
        logger.info(f"✅ Detected {len(detected_questions)} questions (all questions ending with ?) from polished transcript")
        if len(detected_questions) > 0:
            logger.info(f"📋 Sample questions detected: {[q['question'][:50] + '...' if len(q['question']) > 50 else q['question'] for q in detected_questions[:3]]}")
        else:
            logger.warning(f"⚠️ No questions detected. Transcript length: {len(transcript_text)}, Sample text: {transcript_text[:200]}")
        return detected_questions

    async def analyze_interaction_engagement(self, speech_analysis: Dict) -> Dict[str, Any]:
        """
        Analyze instructor-student interaction and questioning techniques
        
        New Logic:
        1. List ALL questions (ending with ? or sounding like questions)
        2. Analyze which questions are high-level pedagogically
        3. Mark high-level questions
        4. Count total questions and high-level questions
        5. Calculate scoring based on questions
        """
        transcript = speech_analysis.get('transcript', '')
        words_data = speech_analysis.get('word_timestamps', [])
        
        # Step 1: Detect ALL questions (ending with ?) - regardless of cognitive level
        all_questions = self.detect_questions_pattern_matching(words_data, transcript)
        
        logger.info(f"📋 Step 1: Found {len(all_questions)} total questions")
        
        # Step 2: Prepare list of all questions for AI analysis
        all_questions_text = ""
        if all_questions:
            all_questions_text = "\n\nAll questions detected (ending with ?):\n"
            for idx, q in enumerate(all_questions, 1):
                timestamp = self.format_timestamp(q['start_time'])
                all_questions_text += f"{idx}. [{timestamp}] {q['question']}\n"
        else:
            all_questions_text = "\n\nNo questions detected (ending with ?)."
        
        context_block = ""
        lc = getattr(self, "lecture_context", None) or ""
        if lc.strip():
            context_block = (
                "\n\n---\nLecture context (provided by the instructor; use when interpreting "
                "questions and teaching intent; reserved for future rubric-based scoring):\n"
                f"{lc.strip()}\n---\n"
            )
        
        # Step 3: AI analysis with ICAP classification (Interactive / Constructive / Active / Passive)
        # ICAP framework (Chi & Wylie): Passive < Active < Constructive < Interactive in cognitive engagement
        response = openai_client.chat.completions.create(
            model="gpt-5-nano",
            messages=[
                {
                    "role": "system",
                    "content": """You are an expert in educational interaction analysis using the ICAP framework (Cognitive Engagement Levels). Classify each question into exactly ONE of these four levels:

ICAP levels (use exactly these labels):
- "Passive": Rhetorical, minimal cognitive demand (e.g., "Right?", "Okay?", "See?"). Expects no real answer.
- "Active": Recall, factual (e.g., "What is photosynthesis?", "Who wrote this?", "When did it happen?"). Single or short factual answer.
- "Constructive": Reasoning, explanation (e.g., "Why does this happen?", "How would you explain...?", "What is the cause?"). Requires building or generating an explanation.
- "Interactive": Co-construction, dialogue (e.g., "Do you agree with her? Why?", "What would you add?", "How does that compare to...?"). Requires building on others' ideas or peer exchange.

Rules:
- Each question must get exactly one label: Passive, Active, Constructive, or Interactive.
- Prefer Constructive over Active if the question invites reasoning (why/how/explain).
- Prefer Interactive if it explicitly references others' ideas, comparison, or building on answers.
- Yes/no or clarification ("Any questions?", "Clear?") are Passive unless they invite elaboration.

Return JSON with:
- all_questions_analyzed: list of {"question": str (exact text), "icap": str ("Passive"|"Active"|"Constructive"|"Interactive")}
- total_questions: int
- count_passive: int
- count_active: int
- count_constructive: int
- count_interactive: int"""
                },
                {
                    "role": "user",
                    "content": f"""Classify each of the following questions from a lecture transcript using ICAP (Passive / Active / Constructive / Interactive). All items below end with a question mark (?).
{context_block}
{all_questions_text}

For each question, set "icap" to exactly one of: Passive, Active, Constructive, Interactive.

Return valid JSON only with: all_questions_analyzed (list of {{"question": "<exact text>", "icap": "<Passive|Active|Constructive|Interactive>"}}), total_questions, count_passive, count_active, count_constructive, count_interactive."""
                }
            ],
            max_completion_tokens=3000
        )
        
        try:
            # Check if response content exists
            response_content = response.choices[0].message.content
            if not response_content:
                raise ValueError("AI response content is None or empty")
            
            analysis = self._safe_json_loads(response_content)
            
            # Step 4: Process AI analysis results (ICAP counts)
            all_questions_analyzed = analysis.get('all_questions_analyzed', [])
            total_questions = analysis.get('total_questions', len(all_questions))
            count_passive = analysis.get('count_passive', 0)
            count_active = analysis.get('count_active', 0)
            count_constructive = analysis.get('count_constructive', 0)
            count_interactive = analysis.get('count_interactive', 0)
            
            # Normalize ICAP from analysis if counts not provided
            if total_questions > 0 and (count_passive + count_active + count_constructive + count_interactive) != total_questions:
                count_passive = sum(1 for q in all_questions_analyzed if (q.get('icap') or '').strip().lower() == 'passive')
                count_active = sum(1 for q in all_questions_analyzed if (q.get('icap') or '').strip().lower() == 'active')
                count_constructive = sum(1 for q in all_questions_analyzed if (q.get('icap') or '').strip().lower() == 'constructive')
                count_interactive = sum(1 for q in all_questions_analyzed if (q.get('icap') or '').strip().lower() == 'interactive')
            
            # Step 5: Match questions to timestamps from pattern-matched questions
            question_timestamp_map = {}
            for q in all_questions:
                question_text_clean = self._normalise_question_key(q['question'])
                question_timestamp_map[question_text_clean] = {
                    'start_time': q['start_time'],
                    'timestamp': self.format_timestamp(q['start_time'])
                }
            
            # Step 6: Build final question list from DETECTED questions (source of truth) + LLM + heuristic ICAP
            final_all_questions = []
            high_level_questions = []  # Constructive + Interactive for backward compatibility
            
            for idx, q_row in enumerate(all_questions):
                question_text = (q_row.get('question') or '').strip()
                if not question_text:
                    continue
                icap_llm = self._find_llm_icap_for_question(all_questions_analyzed, question_text, idx)
                icap_heuristic = self._classify_icap_heuristic(question_text)
                icap = self._merge_icap_labels(icap_llm, icap_heuristic)
                
                question_text_clean = self._normalise_question_key(question_text)
                timestamp_info = question_timestamp_map.get(question_text_clean)
                if timestamp_info:
                    timestamp = timestamp_info['timestamp']
                    start_time = timestamp_info['start_time']
                else:
                    timestamp, start_time = "00:00", 0
                    for q in all_questions:
                        if self._normalise_question_key(question_text)[:28] in self._normalise_question_key(q['question'])[:60] or self._normalise_question_key(q['question'])[:28] in self._normalise_question_key(question_text)[:60]:
                            timestamp = self.format_timestamp(q['start_time'])
                            start_time = q['start_time']
                            break
                
                question_entry = {
                    'question': question_text,
                    'precise_timestamp': timestamp,
                    'start_time': start_time,
                    'icap': icap,
                    'is_high_level': icap in ('Constructive', 'Interactive')
                }
                final_all_questions.append(question_entry)
                if question_entry['is_high_level']:
                    high_level_questions.append(question_entry)

            # Sort questions chronologically for UI readability and for QDS stability
            final_all_questions.sort(key=lambda x: float(x.get('start_time') or 0))
            high_level_questions.sort(key=lambda x: float(x.get('start_time') or 0))
            
            # Recount ICAP from merged labels (do not trust LLM aggregate counts)
            total_questions = len(final_all_questions)
            count_passive = sum(1 for x in final_all_questions if x['icap'] == 'Passive')
            count_active = sum(1 for x in final_all_questions if x['icap'] == 'Active')
            count_constructive = sum(1 for x in final_all_questions if x['icap'] == 'Constructive')
            count_interactive = sum(1 for x in final_all_questions if x['icap'] == 'Interactive')
            
            high_level_questions_count = count_constructive + count_interactive
            
            # Step 7: Quantify metrics (0-10 scale) and percentages (0-100) for 20% sub-category
            # Components: Interaction frequency (QD), Question quality (CLI), Student Uptake Index (SUI), Question Distribution Stability (QDS)
            duration_minutes = max(0.1, speech_analysis.get('duration_minutes', 1))
            effective_minutes = duration_minutes
            duration_seconds = duration_minutes * 60.0
            
            eqd_per_minute = 0.0
            questions_per_minute = 0.0
            qds: Dict[str, Any] = {}
            sui: Dict[str, Any] = {"score": 0.0, "uptake_hits": 0, "uptake_rate": 0.0, "evidence": ""}
            sui_prompt = 0.0
            sui_uptake_raw = 0.0
            sui_evidence = ""
            cli_100 = 0.0
            question_engagement_narrative = ""
            if total_questions == 0:
                interaction_frequency = 0.0
                question_quality = 0.0
                student_uptake_index = 0.0
                question_distribution_stability = 0.0
                cognitive_level = 'low'
                interaction_frequency_pct = 0.0
                question_quality_pct = 0.0
                student_uptake_index_pct = 0.0
                question_distribution_stability_pct = 0.0
                overall_pct = 0.0
            else:
                question_quality, cli_100, cognitive_level = self._compute_question_quality_from_icap_counts(
                    total_questions,
                    count_passive,
                    count_active,
                    count_constructive,
                    count_interactive,
                    effective_minutes,
                )

                # 1) Question density (QD): high 8-10, mid 4-7, low 1-3, no density (0-0.1) = 0
                qd = total_questions / effective_minutes
                if qd <= 0.1:
                    interaction_frequency = 0.0
                elif qd < 0.5:
                    interaction_frequency = 1.0 + (qd - 0.1) / 0.4 * 2.0  # 1 to 3
                elif qd < 1.5:
                    interaction_frequency = 4.0 + (qd - 0.5) / 1.0 * 3.0  # 4 to 7
                else:
                    interaction_frequency = 8.0 + min(2.0, (qd - 1.5) * 2.0)  # 8 to 10 (e.g. 1.5->8, 2.5->10)
                interaction_frequency = round(min(10.0, max(0.0, interaction_frequency)), 1)
                
                # 2) Student Uptake Index (SUI): uptake cues after questions -> 0-10
                eqd = (count_constructive + count_interactive) / effective_minutes
                eqd_per_minute = round(eqd, 3)
                questions_per_minute = round(total_questions / effective_minutes, 3)
                sf_metrics = speech_analysis.get("student_feedback_metrics") or {}
                sui = self._compute_student_uptake_index_from_questions(words_data, all_questions, sf_metrics)
                sui_uptake_raw = float(sui.get("score") or 0.0)
                sui_prompt = self._sui_from_prompting_density(
                    effective_minutes,
                    total_questions,
                    count_active,
                    count_constructive,
                    count_interactive,
                )
                conf_fb = (sf_metrics.get("student_feedback_confidence") or "none").lower()
                if conf_fb in ("none", "low"):
                    student_uptake_index = round(max(sui_uptake_raw, sui_prompt), 1)
                    sui_note = (
                        f" Webcast-style audio: listener uptake is often invisible; SUI uses max(transcript uptake cues, "
                        f"prompting-density proxy {sui_prompt}/10) so high questioning volume is not collapsed to ~2/10."
                    )
                elif conf_fb == "medium":
                    student_uptake_index = round(
                        max(sui_uptake_raw, 0.5 * sui_uptake_raw + 0.5 * sui_prompt),
                        1,
                    )
                    sui_note = f" Blended uptake score with prompting-density proxy ({sui_prompt}/10)."
                else:
                    student_uptake_index = sui_uptake_raw
                    sui_note = ""
                sui_evidence = (sui.get("evidence") or "") + sui_note
                
                # 3) Question Distribution Stability (QDS): spread of questions over lecture -> 0-10
                # Use timestamps from the detected question list (chronological) to avoid entropy collapse when text matching fails.
                qds = self._compute_question_distribution_stability(final_all_questions, duration_seconds)
                question_distribution_stability = float(qds.get("score") or 0.0)

                question_engagement_narrative = self._question_engagement_narrative(
                    total_questions,
                    effective_minutes,
                    count_active,
                    count_interactive,
                    count_constructive,
                    high_level_questions_count,
                )

            # Percentages (0-100) for each component (only recalc when we have questions)
            if total_questions > 0:
                interaction_frequency_pct = round(interaction_frequency * 10.0, 1)
                question_quality_pct = round(question_quality * 10.0, 1)
                student_uptake_index_pct = round(student_uptake_index * 10.0, 1)
                question_distribution_stability_pct = round(question_distribution_stability * 10.0, 1)
                overall_pct = (interaction_frequency + question_quality + student_uptake_index + question_distribution_stability) / 4.0 * 10.0
            
            logger.info(f"✅ Question analysis complete: {total_questions} total | QD→{interaction_frequency} SUI→{student_uptake_index} QDS→{question_distribution_stability}")
            
            return {
                'score': round((interaction_frequency + question_quality + student_uptake_index + question_distribution_stability) / 4.0, 1),
                'interaction_frequency': round(interaction_frequency, 1),
                'question_quality': round(question_quality, 1),
                'student_uptake_index': round(student_uptake_index, 1),
                'student_engagement_opportunities': round(student_uptake_index, 1),  # backward compat
                'question_distribution_stability': round(question_distribution_stability, 1),
                'interaction_frequency_pct': interaction_frequency_pct,
                'question_quality_pct': question_quality_pct,
                'student_uptake_index_pct': student_uptake_index_pct,
                'question_distribution_stability_pct': question_distribution_stability_pct,
                'overall_interaction_pct': round(overall_pct, 1),
                'cognitive_level': cognitive_level,
                'high_level_questions': high_level_questions[:20],
                'all_questions': final_all_questions,
                'total_questions': total_questions,
                'high_level_questions_count': high_level_questions_count,
                'total_interactions': total_questions,
                'icap_counts': {'passive': count_passive, 'active': count_active, 'constructive': count_constructive, 'interactive': count_interactive},
                'cognitive_level_index': round(cli_100, 1) if total_questions else 0,
                'eqd_per_minute': eqd_per_minute,
                'questions_per_minute': questions_per_minute,
                'cli_formula': (
                    "CLI (3.2.1): Interactive share of all questions — >20%→9/10; 10–20%→7/10; 5–<10%→5/10; <5%→3/10; "
                    "small bump if Constructive share is high."
                ),
                'qds_formula': (
                    "QDS (3.2.3): Lecture duration split into 5 equal quintiles (0–20%, 20–40%, …); "
                    "2 points per quintile with ≥1 question (max 10)."
                ),
                'qds_cv': qds.get("cv") if isinstance(qds, dict) else None,
                'qds_mean_gap_seconds': qds.get("mean_gap_s") if isinstance(qds, dict) else None,
                'qds_quintile_hits': qds.get("quintile_hits") if isinstance(qds, dict) else None,
                'qds_quintiles_filled': qds.get("quintiles_filled") if isinstance(qds, dict) else None,
                'sui_uptake_hits': sui.get("uptake_hits") if isinstance(sui, dict) else None,
                'sui_uptake_rate': sui.get("uptake_rate") if isinstance(sui, dict) else None,
                'sui_uptake_raw': sui_uptake_raw if total_questions else None,
                'sui_prompting_proxy': sui_prompt if total_questions else None,
                'sui_evidence': sui_evidence if total_questions else (sui.get("evidence") if isinstance(sui, dict) else None),
                'question_engagement_narrative': question_engagement_narrative,
            }
            
        except (json.JSONDecodeError, ValueError, AttributeError) as e:
            error_msg = str(e) if e else "Unknown error"
            logger.error(f"Error in interaction analysis: {error_msg}")
            # Fallback: pattern-matched questions + heuristic ICAP (avoid all-Active)
            fallback_questions = []
            for q in all_questions:
                qt = (q.get('question') or '').strip()
                icap = self._classify_icap_heuristic(qt)
                fallback_questions.append({
                    'question': q['question'],
                    'precise_timestamp': self.format_timestamp(q['start_time']),
                    'start_time': q['start_time'],
                    'icap': icap,
                    'is_high_level': icap in ('Constructive', 'Interactive')
                })
            total_questions = len(fallback_questions)
            if total_questions == 0:
                return {
                    'score': 0.0,
                    'interaction_frequency': 0.0,
                    'question_quality': 0.0,
                    'student_uptake_index': 0.0,
                    'student_engagement_opportunities': 0.0,
                    'question_distribution_stability': 0.0,
                    'interaction_frequency_pct': 0.0,
                    'question_quality_pct': 0.0,
                    'student_uptake_index_pct': 0.0,
                    'question_distribution_stability_pct': 0.0,
                    'overall_interaction_pct': 0.0,
                    'cognitive_level': 'low',
                    'high_level_questions': [],
                    'all_questions': [],
                    'interaction_moments': [],
                    'total_questions': 0,
                    'total_interactions': 0,
                    'icap_counts': {'passive': 0, 'active': 0, 'constructive': 0, 'interactive': 0},
                    'cognitive_level_index': 0,
                    'eqd_per_minute': 0.0,
                    'questions_per_minute': 0.0,
                    'qds_quintile_hits': [False] * 5,
                    'qds_quintiles_filled': 0,
                }
            dm_fb = max(0.1, speech_analysis.get('duration_minutes', 1))
            ds_fb_sec = dm_fb * 60.0
            qpm_fb = round(total_questions / dm_fb, 3)
            fb_high = [x for x in fallback_questions if x.get('is_high_level')]
            cp = sum(1 for x in fallback_questions if x['icap'] == 'Passive')
            ca = sum(1 for x in fallback_questions if x['icap'] == 'Active')
            cc = sum(1 for x in fallback_questions if x['icap'] == 'Constructive')
            ci = sum(1 for x in fallback_questions if x['icap'] == 'Interactive')
            qq_fb, cli100_fb, cog_fb = self._compute_question_quality_from_icap_counts(
                total_questions, cp, ca, cc, ci, dm_fb
            )
            qd_fb = total_questions / dm_fb
            if qd_fb <= 0.1:
                ifreq_fb = 0.0
            elif qd_fb < 0.5:
                ifreq_fb = 1.0 + (qd_fb - 0.1) / 0.4 * 2.0
            elif qd_fb < 1.5:
                ifreq_fb = 4.0 + (qd_fb - 0.5) / 1.0 * 3.0
            else:
                ifreq_fb = 8.0 + min(2.0, (qd_fb - 1.5) * 2.0)
            ifreq_fb = round(min(10.0, max(0.0, ifreq_fb)), 1)
            eqd_fb = (cc + ci) / dm_fb
            sf_fb = speech_analysis.get("student_feedback_metrics") or {}
            sui0_fb = self._compute_student_uptake_index_from_questions(words_data, all_questions, sf_fb)
            sui_raw_fb = float(sui0_fb.get("score") or 0.0)
            sui_p_fb = self._sui_from_prompting_density(dm_fb, total_questions, ca, cc, ci)
            conf_fbb = (sf_fb.get("student_feedback_confidence") or "none").lower()
            if conf_fbb in ("none", "low"):
                sui_fb = round(max(sui_raw_fb, sui_p_fb), 1)
            elif conf_fbb == "medium":
                sui_fb = round(max(sui_raw_fb, 0.5 * sui_raw_fb + 0.5 * sui_p_fb), 1)
            else:
                sui_fb = sui_raw_fb
            qds_fb = self._compute_question_distribution_stability(fallback_questions, ds_fb_sec)
            qds_sc_fb = float(qds_fb.get("score") or 0.0)
            hl_fb = cc + ci
            narr_fb = self._question_engagement_narrative(total_questions, dm_fb, ca, ci, cc, hl_fb)
            score_fb = round((ifreq_fb + qq_fb + sui_fb + qds_sc_fb) / 4.0, 1)
            oa_fb = score_fb * 10.0
            return {
                'score': score_fb,
                'interaction_frequency': ifreq_fb,
                'question_quality': qq_fb,
                'student_uptake_index': sui_fb,
                'student_engagement_opportunities': sui_fb,
                'question_distribution_stability': qds_sc_fb,
                'interaction_frequency_pct': round(ifreq_fb * 10.0, 1),
                'question_quality_pct': round(qq_fb * 10.0, 1),
                'student_uptake_index_pct': round(sui_fb * 10.0, 1),
                'question_distribution_stability_pct': round(qds_sc_fb * 10.0, 1),
                'overall_interaction_pct': round(oa_fb, 1),
                'cognitive_level': cog_fb,
                'high_level_questions': fb_high[:20],
                'all_questions': fallback_questions,
                'interaction_moments': [],
                'total_questions': total_questions,
                'high_level_questions_count': hl_fb,
                'total_interactions': total_questions,
                'icap_counts': {'passive': cp, 'active': ca, 'constructive': cc, 'interactive': ci},
                'cognitive_level_index': round(cli100_fb, 1),
                'eqd_per_minute': round(eqd_fb, 4),
                'questions_per_minute': qpm_fb,
                'cli_formula': (
                    "CLI (question quality): continuous ICAP mix + higher-order count bonuses "
                    "(fallback path: heuristic ICAP only)."
                ),
                'qds_formula': (
                    "QDS: 5 lecture quintiles × 2 points each when that segment contains ≥1 question."
                ),
                'qds_cv': qds_fb.get("cv"),
                'qds_mean_gap_seconds': qds_fb.get("mean_gap_s"),
                'qds_quintile_hits': qds_fb.get("quintile_hits"),
                'qds_quintiles_filled': qds_fb.get("quintiles_filled"),
                'sui_uptake_hits': sui0_fb.get("uptake_hits"),
                'sui_uptake_rate': sui0_fb.get("uptake_rate"),
                'sui_uptake_raw': sui_raw_fb,
                'sui_prompting_proxy': sui_p_fb,
                'sui_evidence': (sui0_fb.get("evidence") or ""),
                'question_engagement_narrative': narr_fb,
            }
        
    def extract_evidence_from_transcript(self, transcript: str) -> List[str]:
        """
        Extract 1-2 relevant quotes from transcript to support assessment
        """
        if not transcript or len(transcript.strip()) < 50:
            return []
        
        # Split transcript into sentences
        sentences = [s.strip() for s in re.split(r'[.!?]+', transcript) if s.strip() and len(s.strip()) > 20]
        
        if len(sentences) < 2:
            return [transcript[:200] + "..." if len(transcript) > 200 else transcript]
        
        # Look for sentences that might indicate good teaching practices
        evidence_indicators = [
            'example', 'for instance', 'let me explain', 'in other words',
            'first', 'second', 'next', 'finally', 'therefore', 'as a result',
            'important', 'key', 'main', 'primary', 'essential', 'crucial',
            'understand', 'remember', 'note', 'consider', 'think about'
        ]
        
        evidence_sentences = []
        
        # Find sentences with evidence indicators
        for sentence in sentences:
            if any(indicator in sentence.lower() for indicator in evidence_indicators):
                if len(sentence) > 30 and len(sentence) < 150:  # Good length for evidence
                    evidence_sentences.append(sentence)
                    if len(evidence_sentences) >= 2:
                        break
        
        # If we don't have enough evidence sentences, pick the first two meaningful sentences
        if len(evidence_sentences) < 2:
            for sentence in sentences:
                if len(sentence) > 30 and len(sentence) < 150:
                    evidence_sentences.append(sentence)
                    if len(evidence_sentences) >= 2:
                        break
        
        return evidence_sentences[:2]  # Return max 2 pieces of evidence
        
    async def generate_comprehensive_summary(self, speech_analysis: Dict, visual_analysis: Dict, pedagogical_analysis: Dict, interaction_analysis: Dict, overall_score: float, transcript_text: str = "", lecture_context: str = "") -> Dict[str, Any]:
        """
        Generate comprehensive evidence-based summary
        """
        transcript = speech_analysis.get('transcript', transcript_text)
        lc = (lecture_context or "").strip()
        lc_section = ""
        if lc:
            lc_section = f"\n\nInstructor-provided lecture context (subject, topic, learning outcomes, etc.):\n{lc}\n"
        
        # Extract evidence from transcript
        evidence_quotes = self.extract_evidence_from_transcript(transcript)
        
        response = openai_client.chat.completions.create(
            model="gpt-5-nano",
            messages=[
                {
                    "role": "system",
                    "content": f"""You are an expert educational evaluator. Create a comprehensive summary that:
                    1. Reviews teaching content quality and accuracy
                    2. Evaluates presentation effectiveness
                    3. Assesses cognitive skill development
                    4. Uses the provided evidence from transcript to support assessment
                    5. Gives actionable, specific recommendations
                    6. Aligns remarks with any instructor-provided lecture context when relevant
                    
                    Evidence from transcript to use in your assessment: {evidence_quotes}
                    
                    Be specific, evidence-based, and constructive."""
                },
                {
                    "role": "user",
                    "content": f"""Create a comprehensive teaching evaluation summary.
                    
                    Overall Score: {overall_score}/10
                    {lc_section}
                    Speech Analysis: {speech_analysis.get('speaking_rate', 0)} WPM, {speech_analysis.get('filler_ratio', 0)*100:.1f}% filler words
                    Visual Engagement: {visual_analysis.get('scores', {}).get('engagement', 0)}/10
                    Pedagogical Quality: {pedagogical_analysis.get('content_organization', 0)}/10 organization
                    Interaction: {interaction_analysis.get('total_questions', 0)} questions, {interaction_analysis.get('cognitive_level', 'medium')} cognitive level
                    
                    Full Transcript:
                    {transcript}
                    
                    Provide:
                    1. content_review: Assessment of teaching content (2-3 sentences)
                    2. presentation_review: Evaluation of delivery and presentation (2-3 sentences)
                    3. cognitive_skills_review: Analysis of cognitive engagement and skill development (2-3 sentences)
                    4. key_evidence: List of 1-3 specific quotes from transcript with timestamps that exemplify teaching quality
                    5. specific_recommendations: 3-5 highly specific, actionable recommendations
                    6. overall_summary: 1 paragraph comprehensive summary
                    
                    Return as JSON."""
                }
            ],
            max_completion_tokens=1500
        )
        
        try:
            # Check if response content exists
            response_content = response.choices[0].message.content
            if not response_content:
                raise ValueError("AI response content is None or empty")
            
            summary = self._safe_json_loads(response_content)
            return summary
        except (json.JSONDecodeError, ValueError, AttributeError):
            return {
                'content_review': 'The lecture covered the material systematically with clear explanations.',
                'presentation_review': 'The instructor demonstrated good delivery with professional presence.',
                'cognitive_skills_review': 'Students were engaged through various teaching techniques.',
                'key_evidence': [],
                'specific_recommendations': [
                    'Incorporate more interactive elements',
                    'Add visual aids to support complex concepts',
                    'Include regular comprehension checks'
                ],
                'overall_summary': f'This lecture achieved an overall score of {round(overall_score, 1)}/10, demonstrating solid teaching fundamentals with opportunities for enhancement in student interaction and engagement techniques.'
            }
    
    def get_rubric_explanation(self, metric_name: str, value: float, score: float) -> Dict[str, str]:
        """
        Generate explanation for a metric based on the rubric
        Returns rating, justification, and remarks based on score ranges
        """
        def get_score_range(score: float) -> tuple:
            """Determine score range category"""
            if score >= 9.0:
                return (9.0, 10.0, "Excellent")
            elif score >= 7.5:
                return (7.5, 8.9, "Good")
            elif score >= 6.0:
                return (6.0, 7.4, "Average")
            elif score >= 4.0:
                return (4.0, 5.9, "Below Average")
            else:
                return (1.0, 3.9, "Poor")
        
        score_min, score_max, rating = get_score_range(score)
        
        # Speech Analysis Metrics
        if metric_name == "speaking_rate":
            wpm = value
            if score >= 9.0:
                justification = "Optimal cognitive processing speed. Research shows 140-180 WPM allows audience to comprehend complex information while maintaining engagement (Tauroza & Allison, 1990)."
                remarks = "Cultural/linguistic differences (ESL instructors may differ). Context-specific filler words in technical subjects."
            elif score >= 7.5:
                justification = "Acceptable range but may be slightly too slow (less engaging) or fast (harder to process). Still effective for most content."
                remarks = ""
            elif score >= 6.0:
                justification = "Borderline acceptable. Below 120 WPM risks losing attention; above 200 WPM challenges comprehension for complex topics."
                remarks = ""
            elif score >= 4.0:
                justification = "Significantly impacts learning. Too slow = monotonous; too fast = overwhelming."
                remarks = ""
            else:
                justification = "Severely impairs communication effectiveness."
                remarks = ""
                
        elif metric_name == "filler_ratio":
            filler_pct = value * 100
            if score >= 9.0:
                justification = "Professional standard. Minimal distraction, maintains credibility. News anchors average <2%."
                remarks = "Consider subject-specific differences (math vs. humanities)."
            elif score >= 7.5:
                justification = "Noticeable but acceptable. Most experienced teachers fall here."
                remarks = ""
            elif score >= 6.0:
                justification = "Begins to distract. Audience starts noticing \"um\"s and \"uh\"s."
                remarks = ""
            elif score >= 4.0:
                justification = "Significantly impacts credibility. Suggests lack of preparation."
                remarks = ""
            else:
                justification = "Severely distracting. May indicate nervousness or poor content mastery."
                remarks = ""
                
        elif metric_name == "voice_variety":
            index = value
            if score >= 9.0:
                justification = "Highly dynamic. Wide pitch and energy variation maintains attention. Radio/podcast quality."
                remarks = "Subject-specific norms (storytelling vs. equation solving)."
            elif score >= 7.5:
                justification = "Moderate variation. Audience stays engaged most of the time."
                remarks = ""
            elif score >= 6.0:
                justification = "Some variation but risks monotony in longer lectures."
                remarks = ""
            elif score >= 4.0:
                justification = "Limited variation. Students may zone out."
                remarks = ""
            else:
                justification = "Monotone delivery. Severely impacts engagement."
                remarks = ""
                
        elif metric_name == "pause_effectiveness":
            index = value
            if score >= 9.0:
                justification = "Optimal cognitive processing. Strategic pauses allow information processing."
                remarks = "Cultural differences in pause patterns. Context-aware pausing, eg pause after complex concept, or wait for student answer."
            elif score >= 7.5:
                justification = "Good use of pauses but could be more strategic."
                remarks = ""
            elif score >= 6.0:
                justification = "Some pauses but missed opportunities for emphasis."
                remarks = ""
            elif score >= 4.0:
                justification = "Too few (rushed) or too many (hesitant)."
                remarks = ""
            else:
                justification = "Disruptive to flow or completely rushed."
                remarks = ""
                
        elif metric_name == "transcription_confidence":
            confidence_pct = value * 100
            if score >= 9.0:
                justification = "Crystal clear articulation. Whisper transcribes with near-perfect accuracy."
                remarks = "Separate audio quality issues from articulation issues. Account for technical jargon that Whisper may not know. Accent/dialect considerations."
            elif score >= 7.5:
                justification = "Mostly clear with minor unclear moments."
                remarks = ""
            elif score >= 6.0:
                justification = "Noticeable articulation issues or audio quality problems."
                remarks = ""
            elif score >= 4.0:
                justification = "Frequent unclear speech affects understanding."
                remarks = ""
            else:
                justification = "Severely impacts communication. May be audio quality or articulation."
                remarks = ""
        
        # Body Language Metrics
        elif metric_name == "eye_contact":
            if score >= 9.0:
                justification = "Builds connection and trust. Critical for online/recorded lectures. Students report higher engagement with direct eye contact (Chen, 2012)."
                remarks = "Account for legitimate reasons to look away (board work, demonstrations). Cultural sensitivity in eye contact expectations."
            elif score >= 7.5:
                justification = "Strong connection with acceptable reference to notes."
                remarks = ""
            elif score >= 6.0:
                justification = "Connection established but weakened by looking away."
                remarks = ""
            elif score >= 4.0:
                justification = "Weak connection. Appears disengaged or over-reliant on notes."
                remarks = ""
            else:
                justification = "No connection. Reading or distracted."
                remarks = "Limitation/accuracy depends on video recording quality and online lecture materials, ie audience cam not presence."
                
        elif metric_name == "gestures":
            if score >= 9.0:
                justification = "Enhances comprehension. Gestures can improve retention by 12% (Goldin-Meadow, 2000)."
                remarks = "Cultural variations exist. Subject-specific gesture norms (physics vs. literature)."
            elif score >= 7.5:
                justification = "Helpful but could be more intentional."
                remarks = ""
            elif score >= 6.0:
                justification = "Adds some visual interest but limited effectiveness."
                remarks = ""
            elif score >= 4.0:
                justification = "Detracts from message or offers no support."
                remarks = ""
            else:
                justification = "Actively harms communication."
                remarks = ""
                
        elif metric_name == "posture":
            if score >= 9.0:
                justification = "Projects authority and confidence. Students perceive better teacher quality with good posture (Neill & Caswell, 1993)."
                remarks = "Cultural variations exist. Sitting vs standing norms. Movement vs static posture balance."
            elif score >= 7.5:
                justification = "Professional with room for minor improvement."
                remarks = ""
            elif score >= 6.0:
                justification = "Adequate but could project more confidence."
                remarks = ""
            elif score >= 4.0:
                justification = "Undermines authority and engagement."
                remarks = ""
            else:
                justification = "Significantly impacts perceived competence."
                remarks = ""
                
        elif metric_name == "facial_engagement":
            if score >= 9.0:
                justification = "Emotional contagion. Instructor enthusiasm transfers to students (Patrick et al., 2000)."
                remarks = "Content-appropriate expression (serious topic vs. humor). Cultural display rules for emotions. Balance between animated and professional."
            elif score >= 7.5:
                justification = "Engaging with minor flat moments."
                remarks = ""
            elif score >= 6.0:
                justification = "Adequate but could show more energy."
                remarks = ""
            elif score >= 4.0:
                justification = "Fails to convey enthusiasm or appropriate emotion."
                remarks = ""
            else:
                justification = "Disconnects students from content."
                remarks = ""
                
        elif metric_name == "professionalism":
            if score >= 9.0:
                justification = "Establishes credibility. Professional appearance increases perceived expertise (Mack & Rainey, 1990)."
                remarks = "Discipline-specific norms (engineering vs. arts). Cultural/institutional dress codes. Balance formality with approachability."
            elif score >= 7.5:
                justification = "Appropriate for most contexts."
                remarks = ""
            elif score >= 6.0:
                justification = "May reduce perceived authority slightly."
                remarks = ""
            elif score >= 4.0:
                justification = "Detracts from professional image."
                remarks = ""
            else:
                justification = "Seriously undermines credibility."
                remarks = ""
        
        # Teaching Effectiveness Metrics
        elif metric_name == "content_organisation":
            if score >= 9.0:
                justification = "Cognitive load theory. Organised content reduces extraneous load (Sweller, 1988)."
                remarks = "Discipline-specific organisational patterns. Length-appropriate structure expectations."
            elif score >= 7.5:
                justification = "Well-organised with room for refinement."
                remarks = ""
            elif score >= 6.0:
                justification = "Students can follow but may miss connections."
                remarks = ""
            elif score >= 4.0:
                justification = "Students struggle to build mental model."
                remarks = ""
            else:
                justification = "Severely impairs learning."
                remarks = ""
                
        elif metric_name == "engagement_techniques":
            if score >= 9.0:
                justification = "Active learning improves retention by 40-50% (Freeman et al., 2014)."
                remarks = "Quality vs. quantity of engagement. Subject-appropriate engagement types."
            elif score >= 7.5:
                justification = "Engages students but could diversify."
                remarks = ""
            elif score >= 6.0:
                justification = "Primarily lecture format with minimal interaction."
                remarks = ""
            elif score >= 4.0:
                justification = "Students likely passive throughout."
                remarks = ""
            else:
                justification = "Students are passive recipients only."
                remarks = ""
                
        elif metric_name == "communication_clarity":
            if score >= 9.0:
                justification = "Comprehensibility is crucial factor in teaching effectiveness (Feldman, 1976)."
                remarks = ""
            elif score >= 7.5:
                justification = "Effective communication with room for precision."
                remarks = ""
            elif score >= 6.0:
                justification = "Students grasp main ideas but miss details."
                remarks = ""
            elif score >= 4.0:
                justification = "Students struggle to understand concepts."
                remarks = ""
            else:
                justification = "Students cannot follow."
                remarks = ""
                
        elif metric_name == "use_of_examples":
            if score >= 9.0:
                justification = "Concreteness principle. Examples improve transfer by 50% (Gentner et al., 2003)."
                remarks = ""
            elif score >= 7.5:
                justification = "Examples present but could add more variety."
                remarks = ""
            elif score >= 6.0:
                justification = "Basic examples without depth."
                remarks = ""
            elif score >= 4.0:
                justification = "Concepts remain abstract."
                remarks = ""
            else:
                justification = "Students cannot relate to content."
                remarks = ""
                
        elif metric_name == "knowledge_checking":
            if score >= 9.0:
                justification = "Formative assessment doubles learning gains (Black & William, 1998)."
                remarks = ""
            elif score >= 7.5:
                justification = "Monitoring understanding but could be more frequent."
                remarks = ""
            elif score >= 6.0:
                justification = "Assumes understanding without verification."
                remarks = ""
            elif score >= 4.0:
                justification = "May not identify student confusion."
                remarks = ""
            else:
                justification = "No feedback loop for learning."
                remarks = ""
        
        # Interaction & Engagement Metrics
        elif metric_name == "question_frequency":
            if score >= 9.0:
                justification = "Socratic method. Questions drive deep thinking (Overholser, 1992)."
                remarks = ""
            elif score >= 7.5:
                justification = "Regular questioning with room for depth."
                remarks = ""
            elif score >= 6.0:
                justification = "Questions asked but limited cognitive challenge."
                remarks = ""
            elif score >= 4.0:
                justification = "Minimal questioning, low cognitive demand."
                remarks = ""
            else:
                justification = "No inquiry-based learning."
                remarks = ""
                
        elif metric_name == "cognitive_level":
            if score >= 9.0:
                justification = "Critical thinking development. Higher-order thinking improves problem-solving (Zohar & Dori, 2003)."
                remarks = ""
            elif score >= 7.5:
                justification = "Good cognitive challenge with balance."
                remarks = ""
            elif score >= 6.0:
                justification = "Some thinking required but limited depth."
                remarks = ""
            elif score >= 4.0:
                justification = "Primarily surface-level thinking."
                remarks = ""
            else:
                justification = "No critical thinking development."
                remarks = ""
                
        elif metric_name == "interaction_opportunity":
            if score >= 9.0:
                justification = "Student-centered learning. Participation increases ownership (Deci & Ryan, 2000)."
                remarks = ""
            elif score >= 7.5:
                justification = "Students involved but could diversify."
                remarks = ""
            elif score >= 6.0:
                justification = "More instructor-centered than student-centered."
                remarks = ""
            elif score >= 4.0:
                justification = "Primarily passive student role."
                remarks = ""
            else:
                justification = "Students are observers only."
                remarks = ""
        
        elif metric_name == "question_distribution_stability":
            if score >= 9.0:
                justification = "Questions well distributed across the lecture; sustains engagement throughout."
                remarks = ""
            elif score >= 7.5:
                justification = "Good spread of questions with minor clustering."
                remarks = ""
            elif score >= 6.0:
                justification = "Some spread but questions tend to cluster in parts of the lecture."
                remarks = ""
            elif score >= 4.0:
                justification = "Questions concentrated in a limited portion of the session."
                remarks = ""
            else:
                justification = "Questions clustered in one segment or too few to assess distribution."
                remarks = ""
        
        # Presentation Skills Metrics
        elif metric_name == "energy":
            if score >= 9.0:
                justification = "Emotional contagion. Instructor enthusiasm transfers to students (Bettencourt et al., 1983)."
                remarks = ""
            elif score >= 7.5:
                justification = "Engaging with room for more consistency."
                remarks = ""
            elif score >= 6.0:
                justification = "Functional but not motivating."
                remarks = ""
            elif score >= 4.0:
                justification = "Fails to inspire or engage."
                remarks = ""
            else:
                justification = "Disengages students."
                remarks = ""
                
        elif metric_name == "voice_modulation":
            if score >= 9.0:
                justification = "Prosodic emphasis aids comprehension (Hincks, 2005)."
                remarks = ""
            elif score >= 7.5:
                justification = "Engaging vocal delivery."
                remarks = ""
            elif score >= 6.0:
                justification = "Some variety but risks monotony."
                remarks = ""
            elif score >= 4.0:
                justification = "Approaching monotone."
                remarks = ""
            else:
                justification = "Flat, disengaging delivery."
                remarks = ""
                
        elif metric_name == "time_management":
            if score >= 9.0:
                justification = "Pacing affects cognitive load (Sweller, 1988)."
                remarks = ""
            elif score >= 7.5:
                justification = "Effective time use with minor improvements possible."
                remarks = ""
            elif score >= 6.0:
                justification = "Time management needs attention."
                remarks = ""
            elif score >= 4.0:
                justification = "Impacts content coverage and understanding."
                remarks = ""
            else:
                justification = "Disrupts learning severely."
                remarks = ""
        
        else:
            # Default explanation for unknown metrics
            justification = f"This metric scored {score:.1f}/10, indicating {rating.lower()} performance."
            remarks = ""
        
        return {
            'rating': rating,
            'justification': justification,
            'remarks': remarks,
            'score_range': f"{score_min:.1f}-{score_max:.1f}"
        }
    
    async def combine_analysis_enhanced(self, speech_analysis: Dict, visual_analysis: Dict, pedagogical_analysis: Dict, interaction_analysis: Dict, sample_frames: List[Dict] = None) -> Dict[str, Any]:
        """
        Enhanced analysis combination with detailed breakdown and transparency
        """
        # Calculate enhanced component scores
        speech_score = self.calculate_speech_score_enhanced(speech_analysis)
        visual_score = self.calculate_visual_score_enhanced(visual_analysis)
        pedagogy_score = self.calculate_pedagogy_score_enhanced(pedagogical_analysis)
        interaction_score = interaction_analysis.get('score', 7.0)
        presentation_score = (speech_score + visual_score) / 2
        
        # Get category weights from config
        category_weights = ANALYSIS_CONFIG["weights"]
        
        # --- MARS v20260224 overall (Content 20% + Delivery 40% + Engagement 40%) ---
        _content_detail = compute_mars_content_category_score_detailed(pedagogical_analysis)
        mars_content_score = _content_detail["content_category_score"]
        mars_delivery_score = compute_mars_delivery_category_score(speech_score, visual_score)
        mars_engagement_score = compute_mars_engagement_category_score(interaction_analysis)
        mars_overall_score = compute_mars_overall_score(
            mars_content_score, mars_delivery_score, mars_engagement_score
        )

        # Legacy equal-weight five-way score (reference only)
        legacy_score = (
            speech_score * category_weights["speech_analysis"] +
            visual_score * category_weights["body_language"] +
            pedagogy_score * category_weights["teaching_effectiveness"] +
            interaction_score * category_weights["interaction_engagement"] +
            presentation_score * category_weights["presentation_skills"]
        )
        overall_score = round(mars_overall_score, 1)

        _mars_ck = (
            'structural_sequencing', 'logical_consistency', 'closure_framing',
            'conceptual_accuracy', 'causal_reasoning_depth', 'multi_perspective_explanation',
            'example_quality_frequency', 'analogy_concept_bridging', 'representation_diversity',
        )
        _content_criteria_evidence = {
            k: str(pedagogical_analysis.get(f'evidence_{k}', '') or '').strip()
            for k in _mars_ck
        }
        _wpm = float(speech_analysis.get('speaking_rate', 0) or 0)
        _fr_pct = round(float(speech_analysis.get('filler_ratio', 0) or 0) * 100.0, 1)
        _vv = round(float(speech_analysis.get('voice_variety_score', 0) or 0), 3)
        _pe = round(float(speech_analysis.get('pause_effectiveness_score', 0) or 0), 3)
        _tconf = round(float(speech_analysis.get('confidence', 0.8) or 0) * 100.0, 1)
        _fillers = speech_analysis.get('filler_details', []) or []
        _top_fillers = ", ".join([f"{f.get('word')} ({f.get('count')}×)" for f in _fillers[:3] if f.get('word')]) or "none highlighted"
        _wpm_band = (
            "within a typical clarity band"
            if 120 <= _wpm <= 200
            else ("fast for dense content" if _wpm > 200 else "slow (may reduce momentum)")
        )
        _filler_band = (
            "low"
            if _fr_pct <= 2.0
            else ("moderate" if _fr_pct <= 5.0 else ("noticeable" if _fr_pct <= 8.0 else "high"))
        )
        _delivery_speech_evidence = (
            f"Speech category score {round(speech_score, 1)}/10. Evidence from audio metrics: speaking rate {_wpm:.0f} WPM ({_wpm_band}); "
            f"filler ratio {_fr_pct}% ({_filler_band}; top fillers: {_top_fillers}); "
            f"transcription confidence {_tconf}% (higher supports reliable transcript-based scoring). "
            f"Voice variety index {_vv} and pause effectiveness index {_pe} suggest how much prosody and pausing patterns support emphasis and comprehension."
        )

        _vs = visual_analysis.get('scores', {}) or {}

        def _body_metric_float(val, default=None):
            if isinstance(val, (list, tuple)):
                val = val[0] if val else None
            if val is None:
                return default
            try:
                return float(val)
            except (TypeError, ValueError):
                return default

        def _fmt_m10(v):
            x = _body_metric_float(v)
            return f"{x:.1f}" if x is not None else "—"

        _frames_n = int(visual_analysis.get('frames_analyzed', 0) or 0)
        _vs_pairs = {
            "eye contact": _vs.get('eye_contact', None),
            "gestures": _vs.get('gestures', None),
            "posture": _vs.get('posture', None),
            "facial engagement": _vs.get('engagement', None),
            "professionalism": _vs.get('professionalism', None),
        }
        _vs_num = {}
        for k, v in _vs_pairs.items():
            fv = _body_metric_float(v)
            if fv is not None:
                _vs_num[k] = fv
        if _vs_num:
            _best_k = max(_vs_num, key=_vs_num.get)
            _worst_k = min(_vs_num, key=_vs_num.get)
            _best = f"Strongest sub-signal: {_best_k} ({_vs_num[_best_k]:.1f}/10)"
            _worst = f"Weakest sub-signal: {_worst_k} ({_vs_num[_worst_k]:.1f}/10)"
        else:
            _best, _worst = "Strongest sub-signal: —", "Weakest sub-signal: —"
        _delivery_body_evidence = (
            f"Body language category score {round(visual_score, 1)}/10 from {_frames_n} sampled frame(s), same pattern as speech: "
            f"each metric is a vision-model 1–10 score per frame, then temporally aggregated (middle frames weighted slightly higher). "
            f"Sub-metric means — eye contact {_fmt_m10(_vs.get('eye_contact'))}/10; gestures {_fmt_m10(_vs.get('gestures'))}/10; "
            f"posture {_fmt_m10(_vs.get('posture'))}/10; facial engagement {_fmt_m10(_vs.get('engagement'))}/10; "
            f"professionalism {_fmt_m10(_vs.get('professionalism'))}/10. "
            f"{_best}. {_worst}. "
            "Category score is the configured weighted sum of these five means. "
            "Interpret cautiously when the recording hides face or gestures (camera angle, distance, or slide-heavy layouts)."
        )
        
        # Build the result dictionary
        result = {
            'overall_score': round(overall_score, 1),
            'scoring_model': f'MARS_{MARS_RUBRIC_VERSION}',
            'legacy_equal_weight_overall': round(legacy_score, 1),
            'mars_rubric': {
                'version': MARS_RUBRIC_VERSION,
                'main_category_weights': MARS_CONFIG['main_categories'],
                'content_score': round(mars_content_score, 2),
                'content_subscores': {
                    'content_organisation': round(_content_detail['content_organisation_score'], 2),
                    'explanation_quality': round(_content_detail['explanation_quality_score'], 2),
                    'use_of_examples_representation': round(_content_detail['use_of_examples_representation_score'], 2),
                    'content_formula': _content_detail['formula'],
                    'criteria_weights': MARS_CONFIG.get('content_criteria_weights', {}),
                },
                'delivery_score': round(mars_delivery_score, 2),
                'engagement_score': round(mars_engagement_score, 2),
                'formula': '0.20×Content + 0.40×Delivery + 0.40×Engagement',
                'content_criteria': {
                    'structural_sequencing': round(float(pedagogical_analysis.get('structural_sequencing', 7) or 7), 1),
                    'logical_consistency': round(float(pedagogical_analysis.get('logical_consistency', 7) or 7), 1),
                    'closure_framing': round(float(pedagogical_analysis.get('closure_framing', 7) or 7), 1),
                    'conceptual_accuracy': round(float(pedagogical_analysis.get('conceptual_accuracy', 7) or 7), 1),
                    'causal_reasoning_depth': round(float(pedagogical_analysis.get('causal_reasoning_depth', 7) or 7), 1),
                    'multi_perspective_explanation': round(float(pedagogical_analysis.get('multi_perspective_explanation', 7) or 7), 1),
                    'example_quality_frequency': round(float(pedagogical_analysis.get('example_quality_frequency', 7) or 7), 1),
                    'analogy_concept_bridging': round(float(pedagogical_analysis.get('analogy_concept_bridging', 7) or 7), 1),
                    'representation_diversity': round(float(pedagogical_analysis.get('representation_diversity', 7) or 7), 1),
                },
                'content_criteria_evidence': _content_criteria_evidence,
                'delivery_criteria_evidence': {
                    'speech': _delivery_speech_evidence,
                    'body': _delivery_body_evidence,
                },
                'delivery_components': {
                    'speech_analysis_score': round(speech_score, 2),
                    'body_language_score': round(visual_score, 2),
                    'note': 'Delivery = 50% speech category + 50% body language category (each uses five 0–10 sub-metrics per Revised Rubric).',
                },
                'engagement_components': {
                    'question_density': round(float(interaction_analysis.get('interaction_frequency', 0) or 0), 1),
                    'cognitive_level_index_cli': round(float(interaction_analysis.get('question_quality', 0) or 0), 1),
                    'student_uptake_index': round(float(interaction_analysis.get('student_uptake_index', 0) or 0), 1),
                    'question_distribution_stability': round(float(interaction_analysis.get('question_distribution_stability', 0) or 0), 1),
                    'student_question_frequency': round(float(interaction_analysis.get('student_question_frequency_score', 0) or 0), 1),
                    'student_question_cognitive_level': round(float(interaction_analysis.get('student_question_cognitive_score', 0) or 0), 1),
                    'student_feedback_remarks': interaction_analysis.get('student_feedback_remarks', ''),
                },
            },
            
            # Detailed Speech Analysis
            'speech_analysis': {
                'score': round(speech_score, 1),
                'duration_minutes': round(float(speech_analysis.get('duration_minutes', 0) or 0), 2),
                'speaking_rate': round(speech_analysis.get('speaking_rate', 0), 1),
                'clarity': round(10 - (speech_analysis.get('filler_ratio', 0) * 20), 1),
                'pace': round(min(10, max(1, 10 - abs(speech_analysis.get('speaking_rate', 150) - 150) / 20)), 1),
                'confidence': round(speech_analysis.get('confidence', 0.8) * 10, 1),
                'voice_variety': round(speech_analysis.get('voice_variety_score', 0.5) * 10, 1),
                'pause_effectiveness': round(speech_analysis.get('pause_effectiveness_score', 0.5) * 10, 1),
                'feedback': self.generate_speech_feedback_enhanced(speech_analysis),
                'metric_scores': {
                    'speaking_rate': round(min(10, max(1, 10 - abs(speech_analysis.get('speaking_rate', 150) - 150) / 20)), 1),
                    'filler_ratio': round(10 - (speech_analysis.get('filler_ratio', 0) * 20), 1),
                    'voice_variety': round(speech_analysis.get('voice_variety_score', 0.5) * 10, 1),
                    'pause_effectiveness': round(speech_analysis.get('pause_effectiveness_score', 0.5) * 10, 1),
                    'transcription_confidence': round(speech_analysis.get('confidence', 0.8) * 10, 1),
                },
                # Raw metrics
                'raw_metrics': {
                    'total_words': speech_analysis.get('word_count', 0),
                    'duration_minutes': round(speech_analysis.get('duration_minutes', 0), 2),
                    'words_per_minute': round(speech_analysis.get('speaking_rate', 0), 1),
                    'filler_word_count': sum(f['count'] for f in speech_analysis.get('filler_details', [])),
                    'filler_ratio_percentage': round(speech_analysis.get('filler_ratio', 0) * 100, 2),
                    'speaking_time_ratio': round(speech_analysis.get('speaking_ratio', 0.7) * 100, 1),
                    'voice_variety_index': round(speech_analysis.get('voice_variety_score', 0.5), 3),
                    'pause_effectiveness_index': round(speech_analysis.get('pause_effectiveness_score', 0.5), 3),
                    'transcription_confidence': round(speech_analysis.get('confidence', 0.8) * 100, 1)
                },
                # Explanations based on rubric
                'explanations': {
                    'speaking_rate': self.get_rubric_explanation('speaking_rate', speech_analysis.get('speaking_rate', 150), round(min(10, max(1, 10 - abs(speech_analysis.get('speaking_rate', 150) - 150) / 20)), 1)),
                    'filler_ratio': self.get_rubric_explanation('filler_ratio', speech_analysis.get('filler_ratio', 0.05), round(10 - (speech_analysis.get('filler_ratio', 0) * 20), 1)),
                    'voice_variety': self.get_rubric_explanation('voice_variety', speech_analysis.get('voice_variety_score', 0.5), round(speech_analysis.get('voice_variety_score', 0.5) * 10, 1)),
                    'pause_effectiveness': self.get_rubric_explanation('pause_effectiveness', speech_analysis.get('pause_effectiveness_score', 0.5), round(speech_analysis.get('pause_effectiveness_score', 0.5) * 10, 1)),
                    'transcription_confidence': self.get_rubric_explanation('transcription_confidence', speech_analysis.get('confidence', 0.8), round(speech_analysis.get('confidence', 0.8) * 10, 1))
                }
            },
            
            # Detailed Body Language Analysis
            'body_language': {
                'score': round(visual_score, 1),
                'eye_contact': round(visual_analysis.get('scores', {}).get('eye_contact', 7), 1),
                'gestures': round(visual_analysis.get('scores', {}).get('gestures', 7), 1),
                'posture': round(visual_analysis.get('scores', {}).get('posture', 7), 1),
                'engagement': round(visual_analysis.get('scores', {}).get('engagement', 7), 1),
                'professionalism': round(visual_analysis.get('scores', {}).get('professionalism', 8), 1),
                'frames_analyzed': visual_analysis.get('frames_analyzed', 0),
                'feedback': self.generate_visual_feedback_enhanced(visual_analysis),
                'remarks': 'Note: Visual analysis accuracy may be limited if the recording does not adequately capture facial expressions and body gestures due to camera angle, distance, or recording style (e.g., presentation slides with voiceover overlay). Results should be interpreted with consideration of these recording constraints.',
                'metric_scores': {
                    'eye_contact': round(visual_analysis.get('scores', {}).get('eye_contact', 7), 1),
                    'gestures': round(visual_analysis.get('scores', {}).get('gestures', 7), 1),
                    'posture': round(visual_analysis.get('scores', {}).get('posture', 7), 1),
                    'facial_engagement': round(visual_analysis.get('scores', {}).get('engagement', 7), 1),
                    'professionalism': round(visual_analysis.get('scores', {}).get('professionalism', 8), 1),
                },
                # Raw metrics
                'raw_metrics': {
                    'total_frames_extracted': visual_analysis.get('frames_analyzed', 0),
                    'frame_interval_seconds': ANALYSIS_CONFIG["sampling"]["frame_interval_seconds"],
                    'eye_contact_raw': round(visual_analysis.get('scores', {}).get('eye_contact', 7), 2),
                    'gestures_raw': round(visual_analysis.get('scores', {}).get('gestures', 7), 2),
                    'posture_raw': round(visual_analysis.get('scores', {}).get('posture', 7), 2),
                    'engagement_raw': round(visual_analysis.get('scores', {}).get('engagement', 7), 2),
                    'professionalism_raw': round(visual_analysis.get('scores', {}).get('professionalism', 8), 2)
                },
                # Explanations based on rubric
                'explanations': {
                    'eye_contact': self.get_rubric_explanation('eye_contact', visual_analysis.get('scores', {}).get('eye_contact', 7), round(visual_analysis.get('scores', {}).get('eye_contact', 7), 1)),
                    'gestures': self.get_rubric_explanation('gestures', visual_analysis.get('scores', {}).get('gestures', 7), round(visual_analysis.get('scores', {}).get('gestures', 7), 1)),
                    'posture': self.get_rubric_explanation('posture', visual_analysis.get('scores', {}).get('posture', 7), round(visual_analysis.get('scores', {}).get('posture', 7), 1)),
                    'facial_engagement': self.get_rubric_explanation('facial_engagement', visual_analysis.get('scores', {}).get('engagement', 7), round(visual_analysis.get('scores', {}).get('engagement', 7), 1)),
                    'professionalism': self.get_rubric_explanation('professionalism', visual_analysis.get('scores', {}).get('professionalism', 8), round(visual_analysis.get('scores', {}).get('professionalism', 8), 1))
                }
            },
            
            # Detailed Teaching Effectiveness
            'teaching_effectiveness': {
                'score': round(pedagogy_score, 1),
                'mars_content_category_score': round(mars_content_score, 2),
                'content_organization': round(pedagogical_analysis.get('content_organization', 7), 1),
                'engagement_techniques': round(pedagogical_analysis.get('engagement_techniques', 7), 1),
                'communication_clarity': round(pedagogical_analysis.get('communication_clarity', 7), 1),
                'use_of_examples': round(pedagogical_analysis.get('use_of_examples', 7), 1),
                'knowledge_checking': round(pedagogical_analysis.get('knowledge_checking', 7), 1),
                'mars_content_criteria': {
                    'structural_sequencing': round(float(pedagogical_analysis.get('structural_sequencing', 7) or 7), 1),
                    'logical_consistency': round(float(pedagogical_analysis.get('logical_consistency', 7) or 7), 1),
                    'closure_framing': round(float(pedagogical_analysis.get('closure_framing', 7) or 7), 1),
                    'conceptual_accuracy': round(float(pedagogical_analysis.get('conceptual_accuracy', 7) or 7), 1),
                    'causal_reasoning_depth': round(float(pedagogical_analysis.get('causal_reasoning_depth', 7) or 7), 1),
                    'multi_perspective_explanation': round(float(pedagogical_analysis.get('multi_perspective_explanation', 7) or 7), 1),
                    'example_quality_frequency': round(float(pedagogical_analysis.get('example_quality_frequency', 7) or 7), 1),
                    'analogy_concept_bridging': round(float(pedagogical_analysis.get('analogy_concept_bridging', 7) or 7), 1),
                    'representation_diversity': round(float(pedagogical_analysis.get('representation_diversity', 7) or 7), 1),
                },
                'feedback': pedagogical_analysis.get('recommendations', []),
                # Explanations based on rubric
                'explanations': {
                    'content_organisation': self.get_rubric_explanation('content_organisation', pedagogical_analysis.get('content_organization', 7), round(pedagogical_analysis.get('content_organization', 7), 1)),
                    'engagement_techniques': self.get_rubric_explanation('engagement_techniques', pedagogical_analysis.get('engagement_techniques', 7), round(pedagogical_analysis.get('engagement_techniques', 7), 1)),
                    'communication_clarity': self.get_rubric_explanation('communication_clarity', pedagogical_analysis.get('communication_clarity', 7), round(pedagogical_analysis.get('communication_clarity', 7), 1)),
                    'use_of_examples': self.get_rubric_explanation('use_of_examples', pedagogical_analysis.get('use_of_examples', 7), round(pedagogical_analysis.get('use_of_examples', 7), 1)),
                    'knowledge_checking': self.get_rubric_explanation('knowledge_checking', pedagogical_analysis.get('knowledge_checking', 7), round(pedagogical_analysis.get('knowledge_checking', 7), 1))
                }
            },
            
            # Presentation Skills
            'presentation_skills': {
                'score': round(presentation_score, 1),
                'professionalism': round(visual_analysis.get('scores', {}).get('professionalism', 8), 1),
                'energy': round(speech_analysis.get('speaking_ratio', 0.7) * 10, 1),
                'voice_modulation': round(speech_analysis.get('voice_variety_score', 0.5) * 10, 1),
                'time_management': round(self.assess_time_management(speech_analysis), 1),
                'feedback': self.generate_presentation_feedback_enhanced(speech_analysis, visual_analysis),
                # Explanations based on rubric
                'explanations': {
                    'energy': self.get_rubric_explanation('energy', speech_analysis.get('speaking_ratio', 0.7), round(speech_analysis.get('speaking_ratio', 0.7) * 10, 1)),
                    'voice_modulation': self.get_rubric_explanation('voice_modulation', speech_analysis.get('voice_variety_score', 0.5), round(speech_analysis.get('voice_variety_score', 0.5) * 10, 1)),
                    'professionalism': self.get_rubric_explanation('professionalism', visual_analysis.get('scores', {}).get('professionalism', 8), round(visual_analysis.get('scores', {}).get('professionalism', 8), 1)),
                    'time_management': self.get_rubric_explanation('time_management', self.assess_time_management(speech_analysis), round(self.assess_time_management(speech_analysis), 1))
                }
            },

            # NEW: Interaction & Engagement (20% sub-category): QD, Question quality, SUI, QDS → percentages
            'interaction_engagement': {
                'score': round(interaction_score, 1),
                'duration_minutes': round(float(speech_analysis.get('duration_minutes', 0) or 0), 2),
                'interaction_frequency': interaction_analysis.get('interaction_frequency', 0),
                'question_quality': interaction_analysis.get('question_quality', 0),
                'student_uptake_index': interaction_analysis.get('student_uptake_index', interaction_analysis.get('student_engagement_opportunities', 0)),
                'student_engagement_opportunities': interaction_analysis.get('student_uptake_index', interaction_analysis.get('student_engagement_opportunities', 0)),
                'question_distribution_stability': interaction_analysis.get('question_distribution_stability', 0),
                'interaction_frequency_pct': interaction_analysis.get('interaction_frequency_pct', 0),
                'question_quality_pct': interaction_analysis.get('question_quality_pct', 0),
                'student_uptake_index_pct': interaction_analysis.get('student_uptake_index_pct', 0),
                'question_distribution_stability_pct': interaction_analysis.get('question_distribution_stability_pct', 0),
                'overall_interaction_pct': interaction_analysis.get('overall_interaction_pct', 0),
                'student_question_frequency_score': interaction_analysis.get('student_question_frequency_score', 0),
                'student_question_cognitive_score': interaction_analysis.get('student_question_cognitive_score', 0),
                'student_feedback_confidence': interaction_analysis.get('student_feedback_confidence', 'none'),
                'student_feedback_remarks': interaction_analysis.get('student_feedback_remarks', ''),
                'mars_engagement_category_score': round(mars_engagement_score, 2),
                'cognitive_level': interaction_analysis.get('cognitive_level', 'medium'),
                'total_questions': interaction_analysis.get('total_questions', 0),
                'total_interactions': interaction_analysis.get('total_interactions', 0),
                'high_level_questions': interaction_analysis.get('high_level_questions', []),
                'all_questions': interaction_analysis.get('all_questions', []),
                'audience_questions': interaction_analysis.get('audience_questions', []),
                'audience_question_count': interaction_analysis.get('audience_question_count', 0),
                'eqd_per_minute': round(float(interaction_analysis.get('eqd_per_minute', 0) or 0), 4),
                'questions_per_minute': round(float(interaction_analysis.get('questions_per_minute', 0) or 0), 4),
                'icap_counts': interaction_analysis.get('icap_counts', {}),
                'cognitive_level_index': interaction_analysis.get('cognitive_level_index', 0),
                'high_level_questions_count': interaction_analysis.get(
                    'high_level_questions_count',
                    len(interaction_analysis.get('high_level_questions') or []),
                ),
                'question_engagement_narrative': interaction_analysis.get('question_engagement_narrative', ''),
                'sui_prompting_proxy': interaction_analysis.get('sui_prompting_proxy'),
                'sui_uptake_raw': interaction_analysis.get('sui_uptake_raw'),
                'qds_quintile_hits': interaction_analysis.get('qds_quintile_hits'),
                'qds_quintiles_filled': interaction_analysis.get('qds_quintiles_filled'),
                'interaction_moments': interaction_analysis.get('interaction_moments', []),
                'feedback': [
                    f"Asked {interaction_analysis.get('total_questions', 0)} questions at {interaction_analysis.get('cognitive_level', 'medium')} cognitive level",
                    f"Interaction frequency: {interaction_analysis.get('interaction_frequency_pct', 0)}% | Question quality: {interaction_analysis.get('question_quality_pct', 0)}% | SUI: {interaction_analysis.get('student_uptake_index_pct', 0)}% | QDS: {interaction_analysis.get('question_distribution_stability_pct', 0)}%",
                    f"Overall interaction (20% category): {interaction_analysis.get('overall_interaction_pct', 0)}%"
                ],
                # Explanations based on rubric
                'explanations': {
                    'question_frequency': self.get_rubric_explanation('question_frequency', interaction_analysis.get('interaction_frequency', 0), interaction_analysis.get('interaction_frequency', 0)),
                    'cognitive_level': self.get_rubric_explanation('cognitive_level', interaction_analysis.get('question_quality', 0), interaction_analysis.get('question_quality', 0)),
                    'interaction_opportunity': self.get_rubric_explanation('interaction_opportunity', interaction_analysis.get('student_uptake_index', interaction_analysis.get('student_engagement_opportunities', 0)), interaction_analysis.get('student_uptake_index', interaction_analysis.get('student_engagement_opportunities', 0))),
                    'question_distribution_stability': self.get_rubric_explanation('question_distribution_stability', interaction_analysis.get('question_distribution_stability', 0), interaction_analysis.get('question_distribution_stability', 0)),
                }
            },
            
            # Full Transcript with Timecodes
            'full_transcript': {
                'text': speech_analysis.get('transcript', ''),
                'timecoded_words': speech_analysis.get('timecoded_transcript', []),
                'duration_formatted': self.format_timestamp(speech_analysis.get('duration_minutes', 0) * 60),
                'word_count': speech_analysis.get('word_count', 0)
            },
            
            # Filler Words with Timecodes
            'filler_words_detailed': {
                'total_count': sum(f['count'] for f in speech_analysis.get('filler_details', [])),
                'ratio_percentage': round(speech_analysis.get('filler_ratio', 0) * 100, 2),
                'timecoded_occurrences': speech_analysis.get('filler_timecodes', []),
                'breakdown_by_word': speech_analysis.get('filler_details', [])
            },
            
            # Strengths and Improvements
            # Avoid leaking internal parse-failure placeholder strings into end-user reports.
            'strengths': [
                s for s in (pedagogical_analysis.get('strengths', []) or [])
                if isinstance(s, str) and "could not parse" not in s.lower()
            ][:6],
            'improvement_suggestions': [
                s for s in (pedagogical_analysis.get('improvements', []) or [])
                if isinstance(s, str) and "could not parse" not in s.lower()
            ][:6],
            
            # DETAILED CALCULATION BREAKDOWN
            'calculation_breakdown': {
                'mars_primary': {
                    'version': MARS_RUBRIC_VERSION,
                    'overall_score': round(overall_score, 1),
                    'content': round(mars_content_score, 2),
                    'delivery': round(mars_delivery_score, 2),
                    'engagement': round(mars_engagement_score, 2),
                    'formula': '0.20×Content + 0.40×Delivery + 0.40×Engagement',
                    'expanded': (
                        f"0.20×{round(mars_content_score,2)} + 0.40×{round(mars_delivery_score,2)} + 0.40×{round(mars_engagement_score,2)}"
                    ),
                },
                'legacy_reference_equal_weights': {
                    'category_weights': {
                        'speech_analysis': f"{category_weights['speech_analysis']*100}%",
                        'body_language': f"{category_weights['body_language']*100}%",
                        'teaching_effectiveness': f"{category_weights['teaching_effectiveness']*100}%",
                        'interaction_engagement': f"{category_weights['interaction_engagement']*100}%",
                        'presentation_skills': f"{category_weights['presentation_skills']*100}%"
                    },
                    'legacy_overall': round(legacy_score, 2),
                },
                'component_scores': {
                    'speech_analysis': {
                        'score': round(speech_score, 2),
                        'weight': category_weights['speech_analysis'],
                        'contribution': round(speech_score * category_weights['speech_analysis'], 2)
                    },
                    'body_language': {
                        'score': round(visual_score, 2),
                        'weight': category_weights['body_language'],
                        'contribution': round(visual_score * category_weights['body_language'], 2)
                    },
                    'teaching_effectiveness': {
                        'score': round(pedagogy_score, 2),
                        'weight': category_weights['teaching_effectiveness'],
                        'contribution': round(pedagogy_score * category_weights['teaching_effectiveness'], 2)
                    },
                    'interaction_engagement': {
                        'score': round(interaction_score, 2),
                        'weight': category_weights['interaction_engagement'],
                        'contribution': round(interaction_score * category_weights['interaction_engagement'], 2)
                    },
                    'presentation_skills': {
                        'score': round(presentation_score, 2),
                        'weight': category_weights['presentation_skills'],
                        'contribution': round(presentation_score * category_weights['presentation_skills'], 2)
                    }
                },
                'final_calculation': {
                    'formula': 'MARS: 0.20×Content + 0.40×Delivery + 0.40×Engagement',
                    'calculation': (
                        f"0.20×{round(mars_content_score,2)} + 0.40×{round(mars_delivery_score,2)} + 0.40×{round(mars_engagement_score,2)}"
                    ),
                    'result': round(overall_score, 1),
                    'legacy_equal_weight_formula': f'(Speech × {category_weights["speech_analysis"]:.2f}) + ... (see legacy_reference_equal_weights)',
                    'legacy_result': round(legacy_score, 1)
                },
                'speech_breakdown': {
                    'components': category_weights['speech_components'],
                    'scores': {
                        'speaking_rate': round(calculate_metric_score(speech_analysis.get('speaking_rate', 150), SPEECH_METRICS["speaking_rate"]), 2),
                        'clarity': round(calculate_metric_score(speech_analysis.get('confidence', 0.8), SPEECH_METRICS["speaking_clarity"]), 2),
                        'confidence': round(calculate_metric_score(speech_analysis.get('confidence', 0.8), SPEECH_METRICS["speaking_clarity"]), 2),
                        'voice_variety': round(calculate_metric_score(speech_analysis.get('voice_variety_score', 0.5), SPEECH_METRICS["voice_variety"]), 2),
                        'pause_effectiveness': round(calculate_metric_score(speech_analysis.get('pause_effectiveness_score', 0.5), SPEECH_METRICS["pause_effectiveness"]), 2)
                    }
                },
                'visual_breakdown': {
                    'components': category_weights['visual_components'],
                    'scores': visual_analysis.get('scores', {})
                },
                'pedagogy_breakdown': {
                    'components': category_weights['pedagogy_components'],
                    'scores': {
                        'content_organization': pedagogical_analysis.get('content_organization', 7),
                        'engagement_techniques': pedagogical_analysis.get('engagement_techniques', 7),
                        'communication_clarity': pedagogical_analysis.get('communication_clarity', 7),
                        'use_of_examples': pedagogical_analysis.get('use_of_examples', 7),
                        'knowledge_checking': pedagogical_analysis.get('knowledge_checking', 7)
                    }
                }
            },
            
            # Additional Insights
            'detailed_insights': {
                'transcript_summary': speech_analysis.get('transcript', '')[:800] + '...',
                'transcript_length': len(speech_analysis.get('transcript', '')),
                'key_highlights': speech_analysis.get('highlights', [])[:8],
                'visual_observations': visual_analysis.get('observations', [])[:10],
                'filler_word_analysis': speech_analysis.get('filler_details', []),
                'temporal_visual_data': visual_analysis.get('temporal_analysis', {})
            },
            
            # Configuration Used
            'configuration_used': {
                'frames_analyzed': visual_analysis.get('frames_analyzed', 0),
                'frame_interval': ANALYSIS_CONFIG["sampling"]["frame_interval_seconds"],
                'transcript_length': len(speech_analysis.get('transcript', '')),
                'category_weights': category_weights,
                'filler_words_detected': len(speech_analysis.get('filler_details', [])),
                'analysis_timestamp': datetime.now().isoformat()
            }
        }
        
        # Add sample frames for display
        if sample_frames:
            result['sample_frames'] = sample_frames
        
        # Generate Comprehensive Summary AFTER building the result dict
        comprehensive_summary = await self.generate_comprehensive_summary(
            speech_analysis,
            visual_analysis,
            pedagogical_analysis,
            interaction_analysis,
            overall_score,
            speech_analysis.get('transcript', ''),
            lecture_context=getattr(self, "lecture_context", None) or "",
        )
        
        result['comprehensive_summary'] = comprehensive_summary
        
        return result
    
    def calculate_speech_score_enhanced(self, speech_analysis: Dict) -> float:
        """Enhanced speech score calculation with weighted sub-components"""
        weights = ANALYSIS_CONFIG["weights"]["speech_components"]
        
        # Safe float conversion helper
        def safe_float(value, default=0.0):
            if isinstance(value, (list, tuple)):
                value = value[0] if value else default
            try:
                return float(value)
            except (ValueError, TypeError):
                return default
        
        # Calculate individual component scores using configurable thresholds
        speaking_rate = safe_float(speech_analysis.get('speaking_rate', 150))
        rate_score = float(calculate_metric_score(speaking_rate, SPEECH_METRICS["speaking_rate"]))
        
        filler_ratio = safe_float(speech_analysis.get('filler_ratio', 0.05))
        fluency_score = float(calculate_metric_score(filler_ratio, SPEECH_METRICS["filler_ratio"], reverse_scale=True))
        
        confidence = safe_float(speech_analysis.get('confidence', 0.8))
        clarity_score = float(calculate_metric_score(confidence, SPEECH_METRICS["speaking_clarity"]))
        
        voice_variety = safe_float(speech_analysis.get('voice_variety_score', 0.5))
        variety_score = float(calculate_metric_score(voice_variety, SPEECH_METRICS["voice_variety"]))
        
        pause_effectiveness = safe_float(speech_analysis.get('pause_effectiveness_score', 0.5))
        pause_score = float(calculate_metric_score(pause_effectiveness, SPEECH_METRICS["pause_effectiveness"]))
        
        # Weighted combination
        total_score = (
            rate_score * weights["speaking_rate"] +
            fluency_score * weights["clarity"] +
            clarity_score * weights["confidence"] +
            variety_score * weights["voice_variety"] +
            pause_score * weights["pause_effectiveness"]
        )
        
        return float(total_score)
    
    def calculate_visual_score_enhanced(self, visual_analysis: Dict) -> float:
        """Enhanced visual score calculation with weighted sub-components"""
        weights = ANALYSIS_CONFIG["weights"]["visual_components"]
        scores = visual_analysis.get('scores', {})
        
        if not scores:
            return 7.0
        
        # Ensure all scores are floats, not lists or strings
        def safe_float(value, default=7.0):
            if isinstance(value, (list, tuple)):
                value = value[0] if value else default
            try:
                return float(value)
            except (ValueError, TypeError):
                return default
        
        # Calculate weighted visual score with safe conversion
        total_score = (
            safe_float(scores.get('eye_contact', 7)) * weights["eye_contact"] +
            safe_float(scores.get('gestures', 7)) * weights["gestures"] +
            safe_float(scores.get('posture', 7)) * weights["posture"] +
            safe_float(scores.get('engagement', 7)) * weights["engagement"] +
            safe_float(scores.get('professionalism', 8)) * weights["professionalism"]
        )
        
        return float(total_score)
    
    def calculate_pedagogy_score_enhanced(self, pedagogical_analysis: Dict) -> float:
        """Enhanced pedagogy score calculation with weighted sub-components"""
        weights = ANALYSIS_CONFIG["weights"]["pedagogy_components"]
        
        # Use individual component scores instead of just overall_effectiveness
        total_score = (
            pedagogical_analysis.get('content_organization', 7) * weights["content_organization"] +
            pedagogical_analysis.get('engagement_techniques', 7) * weights["engagement_techniques"] +
            pedagogical_analysis.get('communication_clarity', 7) * weights["communication_clarity"] +
            pedagogical_analysis.get('use_of_examples', 7) * weights["use_of_examples"] +
            pedagogical_analysis.get('knowledge_checking', 6.5) * weights["knowledge_checking"]
        )
        
        return total_score
    
    def generate_speech_feedback_enhanced(self, speech_analysis: Dict) -> List[str]:
        """Generate enhanced speech feedback using configurable metrics"""
        feedback = []
        
        speaking_rate = speech_analysis.get('speaking_rate', 150)
        feedback.append(get_metric_feedback(speaking_rate, SPEECH_METRICS["speaking_rate"]))
        
        filler_ratio = speech_analysis.get('filler_ratio', 0.05)
        feedback.append(get_metric_feedback(filler_ratio, SPEECH_METRICS["filler_ratio"], reverse_scale=True))
        
        voice_variety = speech_analysis.get('voice_variety_score', 0.5)
        if voice_variety > 0.7:
            feedback.append("Excellent voice modulation adds engagement and emphasis")
        elif voice_variety < 0.3:
            feedback.append("Consider adding more variation in pitch and energy for better engagement")
        
        pause_effectiveness = speech_analysis.get('pause_effectiveness_score', 0.5)
        if pause_effectiveness > 0.7:
            feedback.append("Strategic use of pauses enhances content delivery")
        elif pause_effectiveness < 0.3:
            feedback.append("Work on using pauses more strategically for emphasis and comprehension")
        
        return feedback
    
    def generate_visual_feedback_enhanced(self, visual_analysis: Dict) -> List[str]:
        """Generate enhanced visual feedback with temporal insights"""
        feedback = []
        scores = visual_analysis.get('scores', {})
        
        # Enhanced feedback based on configurable thresholds
        for metric_name, score in scores.items():
            if metric_name in ['eye_contact', 'gestures', 'posture', 'engagement', 'professionalism']:
                if score >= 8.5:
                    feedback.append(f"Excellent {metric_name.replace('_', ' ')} throughout the session")
                elif score < 6.0:
                    feedback.append(f"Focus on improving {metric_name.replace('_', ' ')} consistency")
        
        # Add temporal analysis feedback if available
        temporal_data = visual_analysis.get('temporal_analysis', {})
        if temporal_data and 'score_progression' in temporal_data:
            progression = temporal_data['score_progression']
            for metric, scores_over_time in progression.items():
                if len(scores_over_time) > 3:
                    trend = np.polyfit(range(len(scores_over_time)), scores_over_time, 1)[0]
                    if trend > 0.1:
                        feedback.append(f"{metric.replace('_', ' ').title()} improved throughout the session")
                    elif trend < -0.1:
                        feedback.append(f"Maintain {metric.replace('_', ' ')} consistency throughout longer sessions")
        
        return feedback
    
    def generate_presentation_feedback_enhanced(self, speech_analysis: Dict, visual_analysis: Dict) -> List[str]:
        """Generate enhanced presentation feedback combining multiple metrics"""
        feedback = []
        
        # Overall energy assessment
        speech_energy = speech_analysis.get('speaking_ratio', 0.7)
        voice_variety = speech_analysis.get('voice_variety_score', 0.5)
        visual_engagement = visual_analysis.get('scores', {}).get('engagement', 7) / 10
        
        overall_energy = (speech_energy + voice_variety + visual_engagement) / 3
        
        if overall_energy > 0.8:
            feedback.append("High energy and dynamic presentation style throughout")
        elif overall_energy < 0.5:
            feedback.append("Consider increasing overall energy and enthusiasm")
        else:
            feedback.append("Good energy balance - maintain consistency")
        
        # Professional delivery assessment
        frames_analyzed = visual_analysis.get('frames_analyzed', 0)
        if frames_analyzed > 20:
            feedback.append(f"Comprehensive analysis of {frames_analyzed} frames shows consistent delivery")
        
        return feedback
    
    
    def assess_time_management(self, speech_analysis: Dict) -> float:
        """Enhanced time management assessment"""
        speaking_ratio = speech_analysis.get('speaking_ratio', 0.7)
        pause_effectiveness = speech_analysis.get('pause_effectiveness_score', 0.5)
        
        # Good time management combines appropriate speaking ratio with effective pauses
        time_score = (speaking_ratio * 0.7 + pause_effectiveness * 0.3) * 10
        
        return min(10.0, max(1.0, time_score))
    
    async def cleanup_temp_files(self, audio_path: Path, video_frames: List):
        """Clean up temporary files"""
        try:
            if audio_path.exists():
                audio_path.unlink()
        except Exception as e:
            print(f"Warning: Could not clean up temporary files: {e}")


def export_questions_to_excel(questions: List[Dict], output_path: Path) -> Optional[Path]:
    """
    Write full question list to an Excel file with columns: #, Question, Timestamp, ICAP.
    ICAP is one of: Interactive, Constructive, Active, Passive.
    Returns output_path if successful, None otherwise.
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available; skipping question Excel export")
        return None
    if not questions:
        return None
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Questions"
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        headers = ["#", "Question", "Timestamp", "ICAP"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 80
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 14
        for row, q in enumerate(questions, 2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=q.get('question', ''))
            ws.cell(row=row, column=3, value=q.get('precise_timestamp', ''))
            icap = q.get('icap', 'Active')
            ws.cell(row=row, column=4, value=icap)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info(f"✅ Exported {len(questions)} questions to {output_path}")
        return output_path
    except Exception as e:
        logger.warning(f"Failed to export questions to Excel: {e}")
        return None


# Global processor instance
video_processor = VideoAnalysisProcessor()